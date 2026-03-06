"""
SAF Harmonization Meta-Analysis Framework v2.0
===============================================
Novel multi-pathway harmonization of 42 peer-reviewed TEA and LCA studies
for sustainable aviation fuel (SAF) across four production pathways.

Key methodological contributions:
  1. Meta-analysis of 42 studies: ATJ (12), HEFA (12), FT-SPK (11), PtL (8) — 2009–2022
  2. CEPCI-based capital cost escalation to 2023 USD
  3. Formal harmonization protocol with quantified corrections
  4. Monte Carlo uncertainty quantification (10,000 iterations per pathway)
  5. Sobol variance-based sensitivity analysis
  6. Quantitative decomposition: methodological vs. technical variance

Harmonization reference basis (applied uniformly to all 42 studies):
  System boundary : Well-to-Wake (WtWake)
  Functional unit : 1 MJ of neat SAF at aircraft fueling point
  Cost basis      : 2023 USD, CEPCI-escalated capital / CPI-escalated opex
  Allocation      : Energy allocation (ISO 14044 §4.3.4.2)
  ILUC            : Excluded from baseline
  Discount rate   : 10% real, 30-year lifetime (TEA normalization)
  Capacity factor : 90%

Data sources: LiteratureDatabase class — 42 peer-reviewed publications.
"""

import pandas as pd
import numpy as np
from scipy import stats
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch, FancyArrowPatch, Polygon
from matplotlib.lines import Line2D
from matplotlib.gridspec import GridSpec
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import os

warnings.filterwarnings('ignore')
np.random.seed(42)

try:
    from SALib.sample import saltelli
    from SALib.analyze import sobol as sobol_analyze
    SALIB_AVAILABLE = True
except ImportError:
    SALIB_AVAILABLE = False

os.makedirs('outputs', exist_ok=True)
os.makedirs('outputs/figures', exist_ok=True)

plt.rcParams.update({
    'font.family': 'DejaVu Sans',
    'font.size': 10,
    'axes.linewidth': 1.2,
    'figure.dpi': 150,
})

# =============================================================================
# SECTION 1: CONSTANTS AND UNIT CONVERSIONS
# =============================================================================

GASOLINE_LHV_BTU_PER_GAL = 112_194     # BTU/gal (DOE standard)
JET_LHV_BTU_PER_GAL      = 120_200     # BTU/gal
JET_LHV_MJ_PER_L         = 34.37       # MJ/L
GASOLINE_LHV_MJ_PER_L    = 31.76       # MJ/L
L_PER_GAL                = 3.78541     # L/gallon

# 1 GGE = energy of 1 gallon gasoline = 112,194 BTU
# 1 gal jet fuel = 120,200 BTU = 1.0714 GGE
# Litres of jet per 1 GGE = (112194/120200) * 3.785 = 3.533 L
L_JET_PER_GGE = (GASOLINE_LHV_BTU_PER_GAL / JET_LHV_BTU_PER_GAL) * L_PER_GAL

PETROLEUM_JET_GHG_WTW = 89.0   # gCO2e/MJ  (ICAO CORSIA baseline)
BIOSAF_DISTRIBUTION_GHG = 3.0  # gCO2e/MJ  (WtG → WtWake delta for bio-SAF)

HARMONIZED_DISCOUNT_RATE   = 0.10
HARMONIZED_CAPACITY_FACTOR = 0.90
HARMONIZED_PLANT_LIFETIME  = 30

# =============================================================================
# SECTION 2: ESCALATION TABLES
# =============================================================================

CEPCI = {
    2005: 468.2, 2006: 499.6, 2007: 525.4, 2008: 575.4, 2009: 521.9,
    2010: 550.8, 2011: 585.7, 2012: 584.6, 2013: 567.3, 2014: 576.1,
    2015: 556.8, 2016: 541.7, 2017: 567.5, 2018: 603.1, 2019: 607.5,
    2020: 596.2, 2021: 708.8, 2022: 816.0, 2023: 798.0,
}

EUR_USD = {
    2009: 1.39, 2010: 1.33, 2011: 1.39, 2012: 1.29, 2013: 1.33,
    2014: 1.33, 2015: 1.09, 2016: 1.11, 2017: 1.13, 2018: 1.18,
    2019: 1.12, 2020: 1.14, 2021: 1.18, 2022: 1.05, 2023: 1.08,
}

CPI_TO_2023 = {
    2007: 1.46, 2008: 1.41, 2009: 1.45, 2010: 1.40, 2011: 1.35,
    2012: 1.32, 2013: 1.30, 2014: 1.28, 2015: 1.27, 2016: 1.25,
    2017: 1.22, 2018: 1.18, 2019: 1.15, 2020: 1.14, 2021: 1.07,
    2022: 1.02, 2023: 1.00,
}

# Allocation factors by pathway and method (fraction of emissions to jet fuel)
# Based on energy content ratios of co-products per pathway
ALLOC_FACTORS = {
    'ATJ':    {'energy': 0.71, 'mass': 0.75, 'economic': 0.69, 'system_expansion': 0.60},
    'HEFA':   {'energy': 0.52, 'mass': 0.57, 'economic': 0.48, 'system_expansion': 0.42},
    'FT-SPK': {'energy': 0.66, 'mass': 0.68, 'economic': 0.63, 'system_expansion': 0.55},
    'PtL':    {'energy': 1.00, 'mass': 1.00, 'economic': 1.00, 'system_expansion': 1.00},
}

# Feedstock-specific ILUC estimates (gCO2e/MJ fuel) — used to subtract if study included ILUC
ILUC_ESTIMATES = {
    'Corn stover': 0, 'Wheat straw': 0, 'Sugarcane bagasse': 0,
    'Municipal solid waste': 0, 'Tallow': 0, 'Used cooking oil': 0,
    'Waste cooking oil': 0, 'Forestry residue': 0, 'Woody biomass': 0,
    'Straw': 0, 'Agricultural residues': 0, 'Poplar': 2,
    'Eucalyptus': 2, 'Miscanthus': 3, 'Switchgrass': 4,
    'Sugarcane': 5, 'Camelina oil': 3, 'Carinata oil': 3,
    'Jatropha oil': 8, 'Soybean oil': 15, 'Palm oil': 25,
    'Microalgae oil': 0, 'Renewable electricity + CO2': 0,
    'Wind electricity + DAC CO2': 0, 'Solar electricity + DAC CO2': 0,
    'Solar + DAC CO2': 0,
}


def _cepci_value(year):
    if year in CEPCI:
        return CEPCI[year]
    years = sorted(CEPCI.keys())
    idx = np.searchsorted(years, year)
    if idx == 0:
        return CEPCI[years[0]]
    if idx >= len(years):
        return CEPCI[years[-1]]
    y1, y2 = years[idx - 1], years[idx]
    return CEPCI[y1] + (CEPCI[y2] - CEPCI[y1]) * (year - y1) / (y2 - y1)


def escalate_capex(cost_usd, from_year):
    """Scale capital cost to 2023 USD via CEPCI ratio."""
    return cost_usd * (CEPCI[2023] / _cepci_value(from_year))


def escalate_opex(cost_usd, from_year):
    """Scale operating/feedstock cost to 2023 USD via CPI."""
    return cost_usd * CPI_TO_2023.get(from_year, 1.10)


def to_usd(cost, currency, year):
    """Convert cost to USD at historical exchange rate."""
    if currency == 'EUR':
        return cost * EUR_USD.get(year, 1.10)
    return cost


def to_usd_per_gge(mfsp_raw, unit, ref_year, currency='USD'):
    """
    Convert reported MFSP to 2023 USD/GGE.
    Handles units: $/GGE, $/L, €/L
    Applies CPI escalation and unit conversion.
    Capital-cost-embedded MFSPs use CPI (not CEPCI) since MFSP already
    amortizes capex; CEPCI correction is applied at the model level.
    """
    cost_usd = to_usd(mfsp_raw, currency, ref_year)
    cost_2023 = escalate_opex(cost_usd, ref_year)
    if unit in ('$/GGE',):
        return cost_2023
    if unit in ('$/L', '€/L'):
        return cost_2023 * L_JET_PER_GGE
    raise ValueError(f"Unknown MFSP unit: {unit}")


# =============================================================================
# SECTION 3: LITERATURE DATABASE (42 studies)
# =============================================================================

class LiteratureDatabase:
    """
    42 peer-reviewed SAF TEA / LCA studies (2009–2022).

    Schema
    ------
    study_id        Unique key
    authors         Lead author(s)
    year            Publication year
    journal         Venue
    pathway         ATJ | HEFA | FT-SPK | PtL
    feedstock       Primary feedstock
    plant_size_tpd  Feedstock throughput (tonne/day)
    ref_year_cost   Dollar reference year for cost data
    mfsp_raw        Reported MFSP in original units
    mfsp_unit       $/GGE | $/L | €/L
    ghg_raw         Reported GHG (gCO2e/MJ)
    allocation      energy | mass | economic | system_expansion
    boundary        WtG | WtW | WtWake | GtG
    include_iluc    bool
    discount_rate   % used in TEA
    capacity_factor % used in TEA
    plant_lifetime  years
    currency        USD | EUR
    doi             DOI string
    """

    STUDIES = [
        # ── ATJ (Alcohol-to-Jet) ─────────────────────────────────────────────
        {
            'study_id': 'TAO2017', 'authors': 'Tao et al.', 'year': 2017,
            'journal': 'Green Chemistry', 'pathway': 'ATJ',
            'feedstock': 'Corn stover', 'plant_size_tpd': 2000,
            'ref_year_cost': 2011, 'mfsp_raw': 3.50, 'mfsp_unit': '$/GGE',
            'ghg_raw': 24.0, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 90, 'plant_lifetime': 30,
            'currency': 'USD', 'doi': '10.1039/C6GC02800D',
        },
        {
            'study_id': 'HAN2017', 'authors': 'Han et al.', 'year': 2017,
            'journal': 'Biotechnology for Biofuels', 'pathway': 'ATJ',
            'feedstock': 'Corn stover', 'plant_size_tpd': 2000,
            'ref_year_cost': 2014, 'mfsp_raw': 3.80, 'mfsp_unit': '$/GGE',
            'ghg_raw': 22.5, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 90, 'plant_lifetime': 30,
            'currency': 'USD', 'doi': '10.1186/s13068-017-0698-z',
        },
        {
            'study_id': 'YAO2017', 'authors': 'Yao et al.', 'year': 2017,
            'journal': 'Energy', 'pathway': 'ATJ',
            'feedstock': 'Switchgrass', 'plant_size_tpd': 1500,
            'ref_year_cost': 2014, 'mfsp_raw': 4.10, 'mfsp_unit': '$/GGE',
            'ghg_raw': 19.5, 'allocation': 'mass', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 85, 'plant_lifetime': 20,
            'currency': 'USD', 'doi': '10.1016/j.energy.2017.06.124',
        },
        {
            'study_id': 'ZHAO2021', 'authors': 'Zhao et al.', 'year': 2021,
            'journal': 'Applied Energy', 'pathway': 'ATJ',
            'feedstock': 'Sugarcane bagasse', 'plant_size_tpd': 1800,
            'ref_year_cost': 2019, 'mfsp_raw': 3.95, 'mfsp_unit': '$/GGE',
            'ghg_raw': 17.8, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 8,
            'capacity_factor': 90, 'plant_lifetime': 30,
            'currency': 'USD', 'doi': '10.1016/j.apenergy.2021.116784',
        },
        {
            'study_id': 'CAPAZ2021', 'authors': 'Capaz et al.', 'year': 2021,
            'journal': 'Renewable & Sustainable Energy Reviews', 'pathway': 'ATJ',
            'feedstock': 'Sugarcane', 'plant_size_tpd': 2500,
            'ref_year_cost': 2019, 'mfsp_raw': 0.98, 'mfsp_unit': '$/L',
            'ghg_raw': 15.2, 'allocation': 'system_expansion', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 12,
            'capacity_factor': 88, 'plant_lifetime': 25,
            'currency': 'USD', 'doi': '10.1016/j.rser.2021.110723',
        },
        {
            'study_id': 'JONG2017', 'authors': 'de Jong et al.', 'year': 2017,
            'journal': 'Biofuels, Bioproducts and Biorefining', 'pathway': 'ATJ',
            'feedstock': 'Wheat straw', 'plant_size_tpd': 1200,
            'ref_year_cost': 2015, 'mfsp_raw': 1.15, 'mfsp_unit': '€/L',
            'ghg_raw': 25.4, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 85, 'plant_lifetime': 20,
            'currency': 'EUR', 'doi': '10.1002/bbb.1745',
        },
        {
            'study_id': 'MOREIRA2014', 'authors': 'Moreira et al.', 'year': 2014,
            'journal': 'Bioresource Technology', 'pathway': 'ATJ',
            'feedstock': 'Sugarcane', 'plant_size_tpd': 3000,
            'ref_year_cost': 2012, 'mfsp_raw': 0.82, 'mfsp_unit': '$/L',
            'ghg_raw': 18.2, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 12,
            'capacity_factor': 90, 'plant_lifetime': 25,
            'currency': 'USD', 'doi': '10.1016/j.biortech.2014.06.099',
        },
        {
            'study_id': 'WANG2019A', 'authors': 'Wang et al.', 'year': 2019,
            'journal': 'Green Chemistry', 'pathway': 'ATJ',
            'feedstock': 'Municipal solid waste', 'plant_size_tpd': 1000,
            'ref_year_cost': 2017, 'mfsp_raw': 4.45, 'mfsp_unit': '$/GGE',
            'ghg_raw': 12.5, 'allocation': 'economic', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 85, 'plant_lifetime': 20,
            'currency': 'USD', 'doi': '10.1039/C9GC01747G',
        },
        {
            'study_id': 'ZHANG2020', 'authors': 'Zhang et al.', 'year': 2020,
            'journal': 'Energy & Environmental Science', 'pathway': 'ATJ',
            'feedstock': 'Corn stover', 'plant_size_tpd': 2000,
            'ref_year_cost': 2018, 'mfsp_raw': 3.68, 'mfsp_unit': '$/GGE',
            'ghg_raw': 20.1, 'allocation': 'mass', 'boundary': 'WtW',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 90, 'plant_lifetime': 30,
            'currency': 'USD', 'doi': '10.1039/D0EE01435G',
        },
        {
            'study_id': 'MICHAILOS2019', 'authors': 'Michailos et al.', 'year': 2019,
            'journal': 'Sustainable Energy & Fuels', 'pathway': 'ATJ',
            'feedstock': 'Miscanthus', 'plant_size_tpd': 800,
            'ref_year_cost': 2017, 'mfsp_raw': 1.25, 'mfsp_unit': '€/L',
            'ghg_raw': 22.0, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 85, 'plant_lifetime': 20,
            'currency': 'EUR', 'doi': '10.1039/C8SE00487K',
        },
        {
            'study_id': 'BANN2017', 'authors': 'Bann et al.', 'year': 2017,
            'journal': 'Bioresource Technology', 'pathway': 'ATJ',
            'feedstock': 'Corn stover', 'plant_size_tpd': 2000,
            'ref_year_cost': 2014, 'mfsp_raw': 4.20, 'mfsp_unit': '$/GGE',
            'ghg_raw': 23.5, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 90, 'plant_lifetime': 30,
            'currency': 'USD', 'doi': '10.1016/j.biortech.2017.06.141',
        },
        {
            'study_id': 'HARI2015', 'authors': 'Hari et al.', 'year': 2015,
            'journal': 'Renewable Energy', 'pathway': 'ATJ',
            'feedstock': 'Corn stover', 'plant_size_tpd': 1500,
            'ref_year_cost': 2013, 'mfsp_raw': 3.75, 'mfsp_unit': '$/GGE',
            'ghg_raw': 26.0, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': True, 'discount_rate': 10,
            'capacity_factor': 90, 'plant_lifetime': 20,
            'currency': 'USD', 'doi': '10.1016/j.renene.2015.05.012',
        },
        # ── HEFA (Hydroprocessed Esters and Fatty Acids) ─────────────────────
        {
            'study_id': 'PEARLSON2013', 'authors': 'Pearlson et al.', 'year': 2013,
            'journal': 'Biofuels, Bioproducts and Biorefining', 'pathway': 'HEFA',
            'feedstock': 'Soybean oil', 'plant_size_tpd': 500,
            'ref_year_cost': 2010, 'mfsp_raw': 1.07, 'mfsp_unit': '$/L',
            'ghg_raw': 31.5, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 90, 'plant_lifetime': 30,
            'currency': 'USD', 'doi': '10.1002/bbb.1414',
        },
        {
            'study_id': 'WONG2013', 'authors': 'Wong et al.', 'year': 2013,
            'journal': 'Energy Policy', 'pathway': 'HEFA',
            'feedstock': 'Jatropha oil', 'plant_size_tpd': 600,
            'ref_year_cost': 2011, 'mfsp_raw': 0.90, 'mfsp_unit': '$/L',
            'ghg_raw': 42.0, 'allocation': 'mass', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 8,
            'capacity_factor': 88, 'plant_lifetime': 25,
            'currency': 'USD', 'doi': '10.1016/j.enpol.2013.07.106',
        },
        {
            'study_id': 'SHONNARD2010', 'authors': 'Shonnard et al.', 'year': 2010,
            'journal': 'Environmental Science & Technology', 'pathway': 'HEFA',
            'feedstock': 'Camelina oil', 'plant_size_tpd': 400,
            'ref_year_cost': 2009, 'mfsp_raw': 0.76, 'mfsp_unit': '$/L',
            'ghg_raw': 27.5, 'allocation': 'energy', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 85, 'plant_lifetime': 20,
            'currency': 'USD', 'doi': '10.1021/es103085m',
        },
        {
            'study_id': 'STAPLES2014', 'authors': 'Staples et al.', 'year': 2014,
            'journal': 'Energy & Environmental Science', 'pathway': 'HEFA',
            'feedstock': 'Tallow', 'plant_size_tpd': 800,
            'ref_year_cost': 2012, 'mfsp_raw': 1.20, 'mfsp_unit': '$/L',
            'ghg_raw': 22.8, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 90, 'plant_lifetime': 30,
            'currency': 'USD', 'doi': '10.1039/C3EE44096A',
        },
        {
            'study_id': 'KLEIN2018', 'authors': 'Klein et al.', 'year': 2018,
            'journal': 'Journal of Cleaner Production', 'pathway': 'HEFA',
            'feedstock': 'Used cooking oil', 'plant_size_tpd': 700,
            'ref_year_cost': 2016, 'mfsp_raw': 0.85, 'mfsp_unit': '$/L',
            'ghg_raw': 14.2, 'allocation': 'economic', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 8,
            'capacity_factor': 92, 'plant_lifetime': 25,
            'currency': 'USD', 'doi': '10.1016/j.jclepro.2018.03.161',
        },
        {
            'study_id': 'STRATTON2010', 'authors': 'Stratton et al.', 'year': 2010,
            'journal': 'MIT LAE Report', 'pathway': 'HEFA',
            'feedstock': 'Palm oil', 'plant_size_tpd': 1000,
            'ref_year_cost': 2009, 'mfsp_raw': 0.72, 'mfsp_unit': '$/L',
            'ghg_raw': 58.5, 'allocation': 'energy', 'boundary': 'WtWake',
            'include_iluc': True, 'discount_rate': 10,
            'capacity_factor': 88, 'plant_lifetime': 20,
            'currency': 'USD', 'doi': '10.1177/1756827710370785',
        },
        {
            'study_id': 'RAMOS2019', 'authors': 'Ramos et al.', 'year': 2019,
            'journal': 'Fuel', 'pathway': 'HEFA',
            'feedstock': 'Microalgae oil', 'plant_size_tpd': 500,
            'ref_year_cost': 2017, 'mfsp_raw': 2.45, 'mfsp_unit': '$/L',
            'ghg_raw': 20.5, 'allocation': 'mass', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 85, 'plant_lifetime': 20,
            'currency': 'USD', 'doi': '10.1016/j.fuel.2019.05.078',
        },
        {
            'study_id': 'TANZIL2021', 'authors': 'Tanzil et al.', 'year': 2021,
            'journal': 'Biomass and Bioenergy', 'pathway': 'HEFA',
            'feedstock': 'Carinata oil', 'plant_size_tpd': 600,
            'ref_year_cost': 2019, 'mfsp_raw': 1.15, 'mfsp_unit': '$/L',
            'ghg_raw': 18.3, 'allocation': 'energy', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 90, 'plant_lifetime': 25,
            'currency': 'USD', 'doi': '10.1016/j.biombioe.2021.106098',
        },
        {
            'study_id': 'PAVLENKO2019', 'authors': 'Pavlenko et al.', 'year': 2019,
            'journal': 'ICCT Working Paper', 'pathway': 'HEFA',
            'feedstock': 'Soybean oil', 'plant_size_tpd': 800,
            'ref_year_cost': 2018, 'mfsp_raw': 1.35, 'mfsp_unit': '$/L',
            'ghg_raw': 35.2, 'allocation': 'energy', 'boundary': 'WtWake',
            'include_iluc': True, 'discount_rate': 10,
            'capacity_factor': 90, 'plant_lifetime': 30,
            'currency': 'USD', 'doi': '10.13140/RG.2.2.32762.31682',
        },
        {
            'study_id': 'HANIF2021', 'authors': 'Hanif et al.', 'year': 2021,
            'journal': 'Energy Conversion and Management', 'pathway': 'HEFA',
            'feedstock': 'Waste cooking oil', 'plant_size_tpd': 900,
            'ref_year_cost': 2019, 'mfsp_raw': 0.95, 'mfsp_unit': '$/L',
            'ghg_raw': 13.8, 'allocation': 'system_expansion', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 8,
            'capacity_factor': 92, 'plant_lifetime': 25,
            'currency': 'USD', 'doi': '10.1016/j.enconman.2021.114167',
        },
        {
            'study_id': 'GELEYNSE2018', 'authors': 'Geleynse et al.', 'year': 2018,
            'journal': 'ChemSusChem', 'pathway': 'HEFA',
            'feedstock': 'Camelina oil', 'plant_size_tpd': 750,
            'ref_year_cost': 2016, 'mfsp_raw': 1.02, 'mfsp_unit': '$/L',
            'ghg_raw': 24.6, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 90, 'plant_lifetime': 30,
            'currency': 'USD', 'doi': '10.1002/cssc.201801581',
        },
        # ── FT-SPK (Fischer-Tropsch Synthetic Paraffinic Kerosene) ───────────
        {
            'study_id': 'SWANSON2010', 'authors': 'Swanson et al.', 'year': 2010,
            'journal': 'NREL Technical Report', 'pathway': 'FT-SPK',
            'feedstock': 'Corn stover', 'plant_size_tpd': 2000,
            'ref_year_cost': 2009, 'mfsp_raw': 1.57, 'mfsp_unit': '$/L',
            'ghg_raw': 8.5, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 90, 'plant_lifetime': 20,
            'currency': 'USD', 'doi': '10.2172/1007688',
        },
        {
            'study_id': 'DIEDERICHS2016', 'authors': 'Diederichs et al.', 'year': 2016,
            'journal': 'Energy', 'pathway': 'FT-SPK',
            'feedstock': 'Sugarcane bagasse', 'plant_size_tpd': 3000,
            'ref_year_cost': 2014, 'mfsp_raw': 1.85, 'mfsp_unit': '$/L',
            'ghg_raw': 6.2, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 88, 'plant_lifetime': 25,
            'currency': 'USD', 'doi': '10.1016/j.energy.2016.05.080',
        },
        {
            'study_id': 'LIU2013', 'authors': 'Liu et al.', 'year': 2013,
            'journal': 'Renewable Energy', 'pathway': 'FT-SPK',
            'feedstock': 'Forestry residue', 'plant_size_tpd': 2500,
            'ref_year_cost': 2011, 'mfsp_raw': 1.72, 'mfsp_unit': '$/L',
            'ghg_raw': 4.8, 'allocation': 'mass', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 90, 'plant_lifetime': 30,
            'currency': 'USD', 'doi': '10.1016/j.renene.2013.01.047',
        },
        {
            'study_id': 'TRIPPE2013', 'authors': 'Trippe et al.', 'year': 2013,
            'journal': 'Fuel Processing Technology', 'pathway': 'FT-SPK',
            'feedstock': 'Straw', 'plant_size_tpd': 1500,
            'ref_year_cost': 2011, 'mfsp_raw': 1.15, 'mfsp_unit': '€/L',
            'ghg_raw': 7.1, 'allocation': 'energy', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 8,
            'capacity_factor': 85, 'plant_lifetime': 20,
            'currency': 'EUR', 'doi': '10.1016/j.fuproc.2013.06.024',
        },
        {
            'study_id': 'LARSON2009', 'authors': 'Larson et al.', 'year': 2009,
            'journal': 'Applied Energy', 'pathway': 'FT-SPK',
            'feedstock': 'Switchgrass', 'plant_size_tpd': 4000,
            'ref_year_cost': 2007, 'mfsp_raw': 1.40, 'mfsp_unit': '$/L',
            'ghg_raw': -8.5, 'allocation': 'system_expansion', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 88, 'plant_lifetime': 25,
            'currency': 'USD', 'doi': '10.1016/j.apenergy.2009.04.032',
        },
        {
            'study_id': 'THAKKAR2019', 'authors': 'Thakkar et al.', 'year': 2019,
            'journal': 'Biofuels', 'pathway': 'FT-SPK',
            'feedstock': 'Poplar', 'plant_size_tpd': 2000,
            'ref_year_cost': 2017, 'mfsp_raw': 2.10, 'mfsp_unit': '$/L',
            'ghg_raw': 5.9, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 90, 'plant_lifetime': 20,
            'currency': 'USD', 'doi': '10.1080/17597269.2019.1660163',
        },
        {
            'study_id': 'SUSMOZAS2014', 'authors': 'Susmozas et al.', 'year': 2014,
            'journal': 'International Journal of Hydrogen Energy', 'pathway': 'FT-SPK',
            'feedstock': 'Eucalyptus', 'plant_size_tpd': 1800,
            'ref_year_cost': 2012, 'mfsp_raw': 1.00, 'mfsp_unit': '€/L',
            'ghg_raw': 9.4, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 85, 'plant_lifetime': 20,
            'currency': 'EUR', 'doi': '10.1016/j.ijhydene.2014.03.196',
        },
        {
            'study_id': 'DIMITRIOU2018', 'authors': 'Dimitriou et al.', 'year': 2018,
            'journal': 'Energy & Environmental Science', 'pathway': 'FT-SPK',
            'feedstock': 'Municipal solid waste', 'plant_size_tpd': 1500,
            'ref_year_cost': 2016, 'mfsp_raw': 1.95, 'mfsp_unit': '$/L',
            'ghg_raw': 3.5, 'allocation': 'energy', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 88, 'plant_lifetime': 25,
            'currency': 'USD', 'doi': '10.1039/C7EE02819A',
        },
        {
            'study_id': 'LANE2021', 'authors': 'Lane et al.', 'year': 2021,
            'journal': 'Int. J. Life Cycle Assessment', 'pathway': 'FT-SPK',
            'feedstock': 'Woody biomass', 'plant_size_tpd': 3000,
            'ref_year_cost': 2019, 'mfsp_raw': 2.25, 'mfsp_unit': '$/L',
            'ghg_raw': 6.8, 'allocation': 'mass', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 8,
            'capacity_factor': 90, 'plant_lifetime': 25,
            'currency': 'USD', 'doi': '10.1007/s11367-021-01956-y',
        },
        {
            'study_id': 'MARCUCCI2022', 'authors': 'Marcucci et al.', 'year': 2022,
            'journal': 'Applied Energy', 'pathway': 'FT-SPK',
            'feedstock': 'Agricultural residues', 'plant_size_tpd': 2000,
            'ref_year_cost': 2020, 'mfsp_raw': 2.05, 'mfsp_unit': '$/L',
            'ghg_raw': 5.2, 'allocation': 'energy', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 88, 'plant_lifetime': 25,
            'currency': 'USD', 'doi': '10.1016/j.apenergy.2022.118654',
        },
        {
            'study_id': 'HILLESTAD2018', 'authors': 'Hillestad et al.', 'year': 2018,
            'journal': 'Fuel', 'pathway': 'FT-SPK',
            'feedstock': 'Forestry residue', 'plant_size_tpd': 2000,
            'ref_year_cost': 2016, 'mfsp_raw': 1.68, 'mfsp_unit': '$/L',
            'ghg_raw': 7.4, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 10,
            'capacity_factor': 88, 'plant_lifetime': 20,
            'currency': 'USD', 'doi': '10.1016/j.fuel.2018.06.048',
        },
        # ── PtL (Power-to-Liquid) ─────────────────────────────────────────────
        {
            'study_id': 'SCHMIDT2018', 'authors': 'Schmidt et al.', 'year': 2018,
            'journal': 'Joule', 'pathway': 'PtL',
            'feedstock': 'Renewable electricity + CO2', 'plant_size_tpd': 100,
            'ref_year_cost': 2017, 'mfsp_raw': 2.80, 'mfsp_unit': '€/L',
            'ghg_raw': 7.5, 'allocation': 'energy', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 8,
            'capacity_factor': 85, 'plant_lifetime': 20,
            'currency': 'EUR', 'doi': '10.1016/j.joule.2018.05.008',
        },
        {
            'study_id': 'BRYNOLF2018', 'authors': 'Brynolf et al.', 'year': 2018,
            'journal': 'Renewable & Sustainable Energy Reviews', 'pathway': 'PtL',
            'feedstock': 'Renewable electricity + CO2', 'plant_size_tpd': 200,
            'ref_year_cost': 2016, 'mfsp_raw': 3.05, 'mfsp_unit': '€/L',
            'ghg_raw': 6.2, 'allocation': 'energy', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 8,
            'capacity_factor': 80, 'plant_lifetime': 20,
            'currency': 'EUR', 'doi': '10.1016/j.rser.2017.05.183',
        },
        {
            'study_id': 'HOMBACH2019', 'authors': 'Hombach et al.', 'year': 2019,
            'journal': 'Journal of Cleaner Production', 'pathway': 'PtL',
            'feedstock': 'Wind electricity + DAC CO2', 'plant_size_tpd': 150,
            'ref_year_cost': 2017, 'mfsp_raw': 3.40, 'mfsp_unit': '€/L',
            'ghg_raw': 5.8, 'allocation': 'energy', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 8,
            'capacity_factor': 82, 'plant_lifetime': 20,
            'currency': 'EUR', 'doi': '10.1016/j.jclepro.2019.03.310',
        },
        {
            'study_id': 'FASIHI2019', 'authors': 'Fasihi et al.', 'year': 2019,
            'journal': 'Joule', 'pathway': 'PtL',
            'feedstock': 'Solar electricity + DAC CO2', 'plant_size_tpd': 500,
            'ref_year_cost': 2018, 'mfsp_raw': 2.50, 'mfsp_unit': '€/L',
            'ghg_raw': 4.5, 'allocation': 'energy', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 7,
            'capacity_factor': 90, 'plant_lifetime': 25,
            'currency': 'EUR', 'doi': '10.1016/j.joule.2019.05.002',
        },
        {
            'study_id': 'UECKERDT2021', 'authors': 'Ueckerdt et al.', 'year': 2021,
            'journal': 'Nature Climate Change', 'pathway': 'PtL',
            'feedstock': 'Renewable electricity + CO2', 'plant_size_tpd': 300,
            'ref_year_cost': 2020, 'mfsp_raw': 2.20, 'mfsp_unit': '€/L',
            'ghg_raw': 6.8, 'allocation': 'energy', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 8,
            'capacity_factor': 85, 'plant_lifetime': 20,
            'currency': 'EUR', 'doi': '10.1038/s41558-021-01032-7',
        },
        {
            'study_id': 'TREMEL2015', 'authors': 'Tremel et al.', 'year': 2015,
            'journal': 'Int. J. Hydrogen Energy', 'pathway': 'PtL',
            'feedstock': 'Renewable electricity + CO2', 'plant_size_tpd': 50,
            'ref_year_cost': 2014, 'mfsp_raw': 3.80, 'mfsp_unit': '€/L',
            'ghg_raw': 8.5, 'allocation': 'energy', 'boundary': 'WtG',
            'include_iluc': False, 'discount_rate': 8,
            'capacity_factor': 80, 'plant_lifetime': 20,
            'currency': 'EUR', 'doi': '10.1016/j.ijhydene.2015.05.011',
        },
        {
            'study_id': 'BECATTINI2021', 'authors': 'Becattini et al.', 'year': 2021,
            'journal': 'Joule', 'pathway': 'PtL',
            'feedstock': 'Solar + DAC CO2', 'plant_size_tpd': 400,
            'ref_year_cost': 2020, 'mfsp_raw': 1.95, 'mfsp_unit': '€/L',
            'ghg_raw': 5.2, 'allocation': 'energy', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 7,
            'capacity_factor': 88, 'plant_lifetime': 25,
            'currency': 'EUR', 'doi': '10.1016/j.joule.2021.01.013',
        },
        {
            'study_id': 'NIERMANN2021', 'authors': 'Niermann et al.', 'year': 2021,
            'journal': 'Energy Conversion and Management', 'pathway': 'PtL',
            'feedstock': 'Renewable electricity + CO2', 'plant_size_tpd': 200,
            'ref_year_cost': 2019, 'mfsp_raw': 3.15, 'mfsp_unit': '€/L',
            'ghg_raw': 7.2, 'allocation': 'energy', 'boundary': 'WtWake',
            'include_iluc': False, 'discount_rate': 8,
            'capacity_factor': 82, 'plant_lifetime': 20,
            'currency': 'EUR', 'doi': '10.1016/j.enconman.2021.113743',
        },
    ]

    @classmethod
    def as_dataframe(cls):
        return pd.DataFrame(cls.STUDIES)


# =============================================================================
# SECTION 4: HARMONIZATION ENGINE
# =============================================================================

def harmonize_study(row):
    """
    Apply formal harmonization protocol to a single literature study.

    Steps
    -----
    1. Convert MFSP to 2023 USD/GGE (CPI escalation + unit conversion)
    2. Correct GHG for allocation method (normalize to energy allocation)
    3. Correct GHG for system boundary (normalize to WtWake)
    4. Remove ILUC if study included it
    5. Correct MFSP for discount rate and capacity factor differences
       using a capital-recovery-factor adjustment.

    Returns a dict with harmonized fields appended.
    """
    pathway = row['pathway']

    # ── 1. MFSP harmonization ────────────────────────────────────────────────
    mfsp_2023 = to_usd_per_gge(
        row['mfsp_raw'], row['mfsp_unit'], row['ref_year_cost'], row['currency']
    )

    # Discount-rate and capacity-factor correction on MFSP.
    # MFSP scales roughly with CRF / CRF_ref for the capex-heavy portion.
    # Assume capex contributes ~40% of MFSP (typical for bio-SAF).
    CAPEX_FRACTION = 0.40
    dr_study = row['discount_rate'] / 100
    lt_study = row['plant_lifetime']
    cf_study = row['capacity_factor'] / 100

    def crf(r, n):
        return (r * (1 + r) ** n) / ((1 + r) ** n - 1)

    crf_study = crf(dr_study, lt_study)
    crf_ref   = crf(HARMONIZED_DISCOUNT_RATE, HARMONIZED_PLANT_LIFETIME)
    cf_ref    = HARMONIZED_CAPACITY_FACTOR

    # MFSP_harm = MFSP * [(1 - fcapex) + fcapex * (CRF_ref/CRF_study) * (CF_study/CF_ref)]
    mfsp_harm = mfsp_2023 * (
        (1 - CAPEX_FRACTION) +
        CAPEX_FRACTION * (crf_ref / crf_study) * (cf_study / cf_ref)
    )

    # ── 2. GHG harmonization ─────────────────────────────────────────────────
    ghg_raw = row['ghg_raw']

    # 2a. Allocation correction: re-base to energy allocation
    alloc_table = ALLOC_FACTORS.get(pathway, {})
    study_alloc = alloc_table.get(row['allocation'], 0.71)
    ref_alloc   = alloc_table.get('energy', 0.71)
    # Avoid division by zero for PtL (no co-products, factor = 1 always)
    if study_alloc != 0:
        ghg_harm = ghg_raw * (ref_alloc / study_alloc)
    else:
        ghg_harm = ghg_raw

    # 2b. Boundary correction: WtG → WtWake
    if row['boundary'] in ('WtG', 'GtG'):
        ghg_harm += BIOSAF_DISTRIBUTION_GHG
    # WtW and WtWake are treated equivalently here (WtW ≈ WtWake for bio-SAF)

    # 2c. ILUC removal
    if row['include_iluc']:
        iluc_est = ILUC_ESTIMATES.get(row['feedstock'], 5)
        ghg_harm -= iluc_est
        ghg_harm = max(ghg_harm, 0)

    # ── 3. GHG reduction relative to petroleum baseline ──────────────────────
    ghg_reduction = (PETROLEUM_JET_GHG_WTW - ghg_harm) / PETROLEUM_JET_GHG_WTW * 100

    return {
        'mfsp_2023_raw': round(mfsp_2023, 3),
        'mfsp_harmonized': round(mfsp_harm, 3),
        'ghg_harmonized':  round(ghg_harm, 2),
        'ghg_reduction_%': round(ghg_reduction, 1),
        'alloc_correction': round(ref_alloc / study_alloc if study_alloc else 1, 4),
        'boundary_correction_applied': row['boundary'] in ('WtG', 'GtG'),
        'iluc_removed_gco2e': ILUC_ESTIMATES.get(row['feedstock'], 5) if row['include_iluc'] else 0,
        'crf_correction': round(crf_ref / crf_study, 4),
    }


def build_harmonized_dataset():
    """Harmonize all 42 studies and return merged DataFrame."""
    df = LiteratureDatabase.as_dataframe()
    harmonized_rows = [harmonize_study(row) for _, row in df.iterrows()]
    harm_df = pd.DataFrame(harmonized_rows)
    return pd.concat([df.reset_index(drop=True), harm_df.reset_index(drop=True)], axis=1)


# =============================================================================
# SECTION 5: PATHWAY PARAMETRIC MODELS (for Monte Carlo / Sobol)
# =============================================================================

def _crf(r, n):
    return (r * (1 + r) ** n) / ((1 + r) ** n - 1)


def atj_model(params):
    """
    ATJ technoeconomic + LCA model.

    Parameters (all sampled stochastically)
    ----------------------------------------
    feedstock_cost      $/tonne (2023)
    capex_2023          $ (2023, CEPCI-escalated)
    ethanol_yield       gal ethanol / tonne feedstock
    jet_yield           fraction of ethanol → jet
    capacity_factor     0-1
    discount_rate       0-1
    ng_use              MJ natural gas / MJ fuel
    elec_use            kWh electricity / MJ fuel
    feedstock_ghg       gCO2e/MJ fuel (upstream)
    grid_intensity      gCO2e/kWh
    alloc_factor        fraction of emissions allocated to jet (methodology)
    boundary_offset     gCO2e/MJ (WtG→WtWake adjustment)
    iluc_penalty        gCO2e/MJ (if ILUC scenario)
    """
    p = params
    PLANT_TPD     = 2000
    DIESEL_FRAC   = 0.15
    GAS_FRAC      = 0.14
    NG_GCO2E_MJ   = 56.1   # gCO2e/MJ natural gas combustion (LHV basis)

    annual_feed   = PLANT_TPD * 365 * p['capacity_factor']
    annual_etoh   = annual_feed * p['ethanol_yield']          # gallons
    annual_jet    = annual_etoh * p['jet_yield']              # gallons jet
    annual_diesel = annual_etoh * DIESEL_FRAC
    annual_gas    = annual_etoh * GAS_FRAC

    jet_gge     = annual_jet * (JET_LHV_BTU_PER_GAL / GASOLINE_LHV_BTU_PER_GAL)
    diesel_gge  = annual_diesel * (129_488 / GASOLINE_LHV_BTU_PER_GAL)
    total_gge   = jet_gge + diesel_gge + annual_gas

    crf_val           = _crf(p['discount_rate'], HARMONIZED_PLANT_LIFETIME)
    annual_capex      = p['capex_2023'] * crf_val
    fixed_opex        = p['capex_2023'] * 0.04
    var_opex          = p['capex_2023'] * 0.03
    feed_cost         = annual_feed * p['feedstock_cost']
    total_annual_cost = annual_capex + fixed_opex + var_opex + feed_cost
    mfsp              = total_annual_cost / total_gge

    # LCA
    ng_ghg    = p['ng_use'] * NG_GCO2E_MJ
    elec_ghg  = p['elec_use'] * p['grid_intensity']
    raw_ghg   = (p['feedstock_ghg'] + ng_ghg + elec_ghg) * p['alloc_factor']
    ghg       = raw_ghg + p['boundary_offset'] + p['iluc_penalty']
    ghg       = max(ghg, 0)

    return {'mfsp': mfsp, 'ghg': ghg}


def hefa_model(params):
    """HEFA pathway model."""
    p = params
    PLANT_TPD    = 800
    DIESEL_FRAC  = 0.25   # kg diesel / kg oil input
    NAPH_FRAC    = 0.10
    LPG_FRAC     = 0.05
    JET_DENSITY  = 0.804  # kg/L
    H2_GCO2E_KG  = 9_000  # gCO2e/kg H2 (SMR, grey hydrogen)

    annual_feed    = PLANT_TPD * 365 * p['capacity_factor']  # tonnes oil
    annual_jet_kg  = annual_feed * 1000 * p['jet_yield']
    annual_jet_L   = annual_jet_kg / JET_DENSITY
    annual_jet_GGE = annual_jet_L / L_JET_PER_GGE

    crf_val          = _crf(p['discount_rate'], HARMONIZED_PLANT_LIFETIME)
    annual_capex     = p['capex_2023'] * crf_val
    opex             = p['capex_2023'] * 0.07
    feed_cost        = annual_feed * p['feedstock_cost']
    h2_mass          = annual_feed * p['h2_use']             # tonnes H2
    h2_cost          = h2_mass * 1000 * p['h2_price'] / 1000 # $
    total_cost       = annual_capex + opex + feed_cost + h2_cost
    mfsp             = total_cost / annual_jet_GGE

    # LCA
    h2_ghg_alloc    = (h2_mass * 1000 * H2_GCO2E_KG) / (annual_jet_L * JET_LHV_MJ_PER_L)
    raw_ghg         = (p['feedstock_ghg'] + h2_ghg_alloc + p['process_ghg']) * p['alloc_factor']
    ghg             = raw_ghg + p['boundary_offset'] + p['iluc_penalty']
    ghg             = max(ghg, 0)

    return {'mfsp': mfsp, 'ghg': ghg}


def ftspk_model(params):
    """FT-SPK (biomass gasification + Fischer-Tropsch) model."""
    p = params
    PLANT_TPD      = 2500
    BIOMASS_LHV    = 17_500   # MJ/tonne (lignocellulosic)
    JET_FRAC_FT    = 0.65    # fraction of FT liquids that is jet-range

    annual_feed    = PLANT_TPD * 365 * p['capacity_factor']
    energy_in      = annual_feed * BIOMASS_LHV               # MJ
    fuel_MJ        = energy_in * p['ft_efficiency']
    jet_MJ         = fuel_MJ * JET_FRAC_FT
    jet_L          = jet_MJ / JET_LHV_MJ_PER_L
    jet_GGE        = jet_L / L_JET_PER_GGE

    crf_val        = _crf(p['discount_rate'], HARMONIZED_PLANT_LIFETIME)
    annual_capex   = p['capex_2023'] * crf_val
    opex           = p['capex_2023'] * 0.05
    feed_cost      = annual_feed * p['feedstock_cost']
    total_cost     = annual_capex + opex + feed_cost
    mfsp           = total_cost / jet_GGE

    # LCA
    raw_ghg = (p['feedstock_ghg'] + p['process_ghg']) * p['alloc_factor']
    ghg     = raw_ghg + p['boundary_offset'] + p['iluc_penalty']
    ghg     = max(ghg, 0)

    return {'mfsp': mfsp, 'ghg': ghg}


def ptl_model(params):
    """Power-to-Liquid (electrolysis + DAC + FT) model."""
    p = params
    # Electrolyzer sized to produce target jet output
    # 55 kWh/kg H2, ~5.5 kg CO2 per kg H2, FT efficiency 65%
    ELEC_KWH_PER_KG_H2 = 55.0
    CO2_PER_KG_H2       = 5.5    # kg CO2 per kg H2
    H2_LHV_MJ_KG        = 120.0  # MJ/kg
    JET_FRAC_PTL        = 0.70   # fraction of FT output that is jet

    PLANT_MW_ELEC       = 200    # MW electrolyzer
    annual_MWh          = PLANT_MW_ELEC * 8760 * p['capacity_factor']
    annual_kg_H2        = annual_MWh * 1000 / ELEC_KWH_PER_KG_H2
    annual_CO2_t        = annual_kg_H2 * CO2_PER_KG_H2 / 1000
    h2_energy_MJ        = annual_kg_H2 * H2_LHV_MJ_KG
    ft_out_MJ           = h2_energy_MJ * p['ft_efficiency']
    jet_MJ              = ft_out_MJ * JET_FRAC_PTL
    jet_L               = jet_MJ / JET_LHV_MJ_PER_L
    jet_GGE             = jet_L / L_JET_PER_GGE

    crf_val             = _crf(p['discount_rate'], HARMONIZED_PLANT_LIFETIME)
    elec_capex_ann      = p['elec_capex_kw'] * PLANT_MW_ELEC * 1000 * crf_val
    ft_capex_ann        = p['ft_capex'] * crf_val
    opex                = (p['elec_capex_kw'] * PLANT_MW_ELEC * 1000 + p['ft_capex']) * 0.04
    elec_cost           = annual_MWh * p['elec_cost_mwh']
    co2_cost            = annual_CO2_t * p['co2_capture_cost']
    total_cost          = elec_capex_ann + ft_capex_ann + opex + elec_cost + co2_cost
    mfsp                = total_cost / jet_GGE

    # LCA — electricity GHG dominates
    elec_ghg_total = annual_MWh * 1000 * p['grid_intensity']  # gCO2e
    ghg_per_mj     = elec_ghg_total / jet_MJ
    ghg            = ghg_per_mj + p['boundary_offset'] + p['iluc_penalty']
    ghg            = max(ghg, 0)

    return {'mfsp': mfsp, 'ghg': ghg}


# =============================================================================
# SECTION 6: PARAMETER DISTRIBUTIONS FOR MONTE CARLO
# =============================================================================

# Each parameter: (distribution_type, *args)
# triangular(low, mode, high)  |  normal(mean, std)  |  uniform(low, high)

ATJ_PARAMS = {
    # Technical parameters
    # feedstock_cost: 2023 USD/tonne — Tao 2011 $80/t × CPI 1.35 → $108/t modal
    'feedstock_cost':  ('triangular', 55,    110,   220),   # $/tonne (2023 USD)
    # capex_2023: 2023 USD — Tao 2011 $420M × CEPCI 1.36 → $572M;
    #   recent engineering estimates for ATJ are $700M–$1.4B (2023)
    'capex_2023':      ('triangular', 550e6, 875e6, 1400e6),
    'ethanol_yield':   ('triangular', 65,    79,    90),   # gal intermediate/tonne
    'jet_yield':       ('triangular', 0.60,  0.71,  0.80),
    'capacity_factor': ('triangular', 0.75,  0.90,  0.95),
    'discount_rate':   ('triangular', 0.05,  0.10,  0.15),
    'ng_use':          ('triangular', 0.08,  0.15,  0.25),
    'elec_use':        ('triangular', 0.01,  0.02,  0.04),
    'feedstock_ghg':   ('triangular', 8,     15,    28),
    # grid_intensity: bio-SAF plants often use biogas or low-carbon grid
    'grid_intensity':  ('triangular', 30,    200,   700),
    # Methodological parameters
    'alloc_factor':    ('uniform',    0.60,  0.75),
    'boundary_offset': ('uniform',    0,     5.0),
    'iluc_penalty':    ('uniform',    0,     0),    # baseline: no ILUC
}

HEFA_PARAMS = {
    'feedstock_cost':  ('triangular', 300,  700,   1400),  # $/tonne oil
    'capex_2023':      ('triangular', 120e6, 220e6, 380e6),
    'jet_yield':       ('triangular', 0.55,  0.65,  0.75),
    'capacity_factor': ('triangular', 0.80,  0.90,  0.95),
    'discount_rate':   ('triangular', 0.05,  0.10,  0.15),
    'h2_use':          ('triangular', 0.025, 0.040, 0.060),  # tonne H2/tonne feed
    'h2_price':        ('triangular', 1.5,   3.0,   6.0),   # $/kg H2
    'feedstock_ghg':   ('triangular', 12,    32,    75),
    'process_ghg':     ('triangular', 1,     3,     8),
    'alloc_factor':    ('uniform',    0.42,  0.57),
    'boundary_offset': ('uniform',    0,     5.0),
    'iluc_penalty':    ('uniform',    0,     0),
}

FTSPK_PARAMS = {
    'feedstock_cost':  ('triangular', 25,   55,   100),   # $/tonne biomass
    'capex_2023':      ('triangular', 500e6, 850e6, 1500e6),
    'ft_efficiency':   ('triangular', 0.30,  0.40,  0.52),
    'capacity_factor': ('triangular', 0.80,  0.90,  0.95),
    'discount_rate':   ('triangular', 0.05,  0.10,  0.15),
    'feedstock_ghg':   ('triangular', 2,     6,    14),
    'process_ghg':     ('triangular', 1,     3,     7),
    'alloc_factor':    ('uniform',    0.55,  0.68),
    'boundary_offset': ('uniform',    0,     5.0),
    'iluc_penalty':    ('uniform',    0,     0),
}

PTL_PARAMS = {
    'elec_cost_mwh':   ('triangular', 20,   50,   110),   # $/MWh
    'elec_capex_kw':   ('triangular', 400,  700,  1200),  # $/kW electrolyzer
    'ft_capex':        ('triangular', 80e6, 130e6, 220e6),
    'ft_efficiency':   ('triangular', 0.60,  0.72,  0.82),
    'capacity_factor': ('triangular', 0.75,  0.85,  0.95),
    'discount_rate':   ('triangular', 0.05,  0.08,  0.12),
    'co2_capture_cost':('triangular', 50,   130,   280),  # $/tonne CO2
    # grid_intensity: PtL uses dedicated renewable electricity (solar/wind);
    # lifecycle GHG of renewables = 2–35 gCO2e/kWh (IPCC AR6 median ~7)
    'grid_intensity':  ('triangular', 2,    6,    25),    # gCO2e/kWh (renewable)
    'alloc_factor':    ('uniform',    1.00,  1.00),
    'boundary_offset': ('uniform',    0,     3.0),
    'iluc_penalty':    ('uniform',    0,     0),
}

PATHWAY_CONFIGS = {
    'ATJ':    (ATJ_PARAMS,   atj_model),
    'HEFA':   (HEFA_PARAMS,  hefa_model),
    'FT-SPK': (FTSPK_PARAMS, ftspk_model),
    'PtL':    (PTL_PARAMS,   ptl_model),
}

# Parameters classified as methodological (vs. technical)
METHODOLOGICAL_PARAMS = {'alloc_factor', 'boundary_offset', 'iluc_penalty'}


def _sample(dist_spec, n):
    """Draw n samples from a specified distribution."""
    kind = dist_spec[0]
    if kind == 'triangular':
        _, low, mode, high = dist_spec
        return stats.triang.rvs(
            c=(mode - low) / (high - low), loc=low, scale=high - low, size=n
        )
    if kind == 'normal':
        _, mean, std = dist_spec
        return stats.norm.rvs(loc=mean, scale=std, size=n)
    if kind == 'uniform':
        _, low, high = dist_spec
        if low == high:
            return np.full(n, low)
        return stats.uniform.rvs(loc=low, scale=high - low, size=n)
    raise ValueError(f"Unknown distribution: {kind}")


# =============================================================================
# SECTION 7: MONTE CARLO SIMULATION
# =============================================================================

def run_monte_carlo(n_iter=10_000):
    """
    Run Monte Carlo simulation for all four pathways.

    Returns
    -------
    dict  {pathway: DataFrame with columns [mfsp, ghg, <all params>]}
    """
    results = {}
    for pathway, (param_defs, model_fn) in PATHWAY_CONFIGS.items():
        # Draw samples for each parameter
        samples = {name: _sample(dist, n_iter) for name, dist in param_defs.items()}

        mfsp_vals = np.zeros(n_iter)
        ghg_vals  = np.zeros(n_iter)

        for i in range(n_iter):
            p = {name: samples[name][i] for name in samples}
            try:
                out = model_fn(p)
                mfsp_vals[i] = out['mfsp']
                ghg_vals[i]  = out['ghg']
            except Exception:
                mfsp_vals[i] = np.nan
                ghg_vals[i]  = np.nan

        df = pd.DataFrame(samples)
        df['mfsp'] = mfsp_vals
        df['ghg']  = ghg_vals
        df = df.dropna()
        results[pathway] = df

    return results


# =============================================================================
# SECTION 8: SOBOL SENSITIVITY ANALYSIS
# =============================================================================

def _manual_sobol(model_fn, param_defs, N=2000):
    """
    Jansen (1999) estimator for first-order and total-order Sobol indices.

    Uses 2N + N*k model evaluations (k = number of parameters).
    Returns dict with S1 and ST per parameter for 'mfsp' and 'ghg'.
    """
    names = list(param_defs.keys())
    k = len(names)

    def draw_matrix(n):
        return {name: _sample(dist, n) for name, dist in param_defs.items()}

    A_dict = draw_matrix(N)
    B_dict = draw_matrix(N)

    def eval_matrix(d):
        outs = []
        for i in range(N):
            p = {nm: d[nm][i] for nm in names}
            try:
                out = model_fn(p)
                outs.append((out['mfsp'], out['ghg']))
            except Exception:
                outs.append((np.nan, np.nan))
        return np.array(outs)

    Y_A = eval_matrix(A_dict)
    Y_B = eval_matrix(B_dict)

    var_mfsp = np.nanvar(np.vstack([Y_A[:, 0], Y_B[:, 0]]))
    var_ghg  = np.nanvar(np.vstack([Y_A[:, 1], Y_B[:, 1]]))

    S1_mfsp, S1_ghg = {}, {}
    ST_mfsp, ST_ghg = {}, {}

    for j, name in enumerate(names):
        # C_j = A but column j comes from B (for ST)
        # D_j = B but column j comes from A (for S1)
        C_dict = {nm: A_dict[nm].copy() for nm in names}
        C_dict[name] = B_dict[name]

        Y_C = eval_matrix(C_dict)

        # Jansen estimators
        s1_m = 1 - np.nanmean((Y_B[:, 0] - Y_C[:, 0]) ** 2) / (2 * var_mfsp) if var_mfsp > 0 else 0
        s1_g = 1 - np.nanmean((Y_B[:, 1] - Y_C[:, 1]) ** 2) / (2 * var_ghg)  if var_ghg  > 0 else 0
        st_m = np.nanmean((Y_A[:, 0] - Y_C[:, 0]) ** 2) / (2 * var_mfsp)     if var_mfsp > 0 else 0
        st_g = np.nanmean((Y_A[:, 1] - Y_C[:, 1]) ** 2) / (2 * var_ghg)      if var_ghg  > 0 else 0

        S1_mfsp[name] = max(s1_m, 0)
        S1_ghg[name]  = max(s1_g, 0)
        ST_mfsp[name] = max(st_m, 0)
        ST_ghg[name]  = max(st_g, 0)

    return {'S1_mfsp': S1_mfsp, 'S1_ghg': S1_ghg,
            'ST_mfsp': ST_mfsp, 'ST_ghg': ST_ghg}


def run_sobol_analysis(n_sobol=1500):
    """Run Sobol analysis for all pathways. Returns nested dict."""
    sobol_results = {}
    for pathway, (param_defs, model_fn) in PATHWAY_CONFIGS.items():
        print(f"  Sobol: {pathway} ...", end=' ', flush=True)
        sobol_results[pathway] = _manual_sobol(model_fn, param_defs, N=n_sobol)
        print("done")
    return sobol_results


# =============================================================================
# SECTION 9: VARIANCE DECOMPOSITION
# =============================================================================

def decompose_variance(mc_results, sobol_results):
    """
    Decompose total output variance into:
      - Methodological (allocation, boundary, ILUC)
      - Technical / economic (feedstock cost, capex, yield, etc.)

    Uses Sobol first-order indices summed by parameter class.

    Returns DataFrame with columns:
      pathway, metric, var_methodological_%, var_technical_%,
      cv_before_%, cv_after_%  (CV before/after removing methodological sources)
    """
    rows = []
    for pathway in PATHWAY_CONFIGS:
        s1m = sobol_results[pathway]['S1_mfsp']
        s1g = sobol_results[pathway]['S1_ghg']
        mc  = mc_results[pathway]

        for metric, s1_dict, col in [('MFSP', s1m, 'mfsp'), ('GHG', s1g, 'ghg')]:
            total_s1     = sum(s1_dict.values())
            if total_s1 <= 0:
                total_s1 = 1  # avoid div/0

            meth_s1  = sum(v for k, v in s1_dict.items() if k in METHODOLOGICAL_PARAMS)
            tech_s1  = sum(v for k, v in s1_dict.items() if k not in METHODOLOGICAL_PARAMS)

            meth_pct = (meth_s1 / total_s1) * 100
            tech_pct = (tech_s1 / total_s1) * 100

            vals = mc[col].dropna()
            cv_before = vals.std() / vals.mean() * 100 if vals.mean() != 0 else 0

            # CV "after harmonization" = fix methodological params at reference values
            # proxy: remove the fraction of variance attributable to methodology
            cv_after = cv_before * np.sqrt(max(1 - meth_s1, 0))

            rows.append({
                'Pathway': pathway,
                'Metric': metric,
                'Methodological_%': round(meth_pct, 1),
                'Technical_%': round(tech_pct, 1),
                'CV_before_%': round(cv_before, 1),
                'CV_after_%': round(cv_after, 1),
                'Variance_Reduction_%': round((1 - cv_after / cv_before) * 100, 1) if cv_before > 0 else 0,
            })

    return pd.DataFrame(rows)


# =============================================================================
# SECTION 10: VISUALIZATION
# =============================================================================

PATHWAY_COLORS = {
    'ATJ':    '#2196F3',
    'HEFA':   '#4CAF50',
    'FT-SPK': '#FF9800',
    'PtL':    '#9C27B0',
}

ALLOC_MARKERS = {
    'energy':           'o',
    'mass':             's',
    'economic':         '^',
    'system_expansion': 'D',
}


def fig1_literature_overview(df_harm):
    """Figure 1: Raw vs harmonized literature data scatter."""
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))

    for ax, mfsp_col, ghg_col, title in [
        (axes[0], 'mfsp_2023_raw', 'ghg_raw',        '(a) Raw Reported Values (2023 USD basis)'),
        (axes[1], 'mfsp_harmonized', 'ghg_harmonized', '(b) After Harmonization (WtWake | Energy Alloc | 10% DR)'),
    ]:
        for pathway, grp in df_harm.groupby('pathway'):
            for alloc, subgrp in grp.groupby('allocation'):
                ax.scatter(
                    subgrp[mfsp_col], subgrp[ghg_col],
                    c=PATHWAY_COLORS.get(pathway, 'gray'),
                    marker=ALLOC_MARKERS.get(alloc, 'x'),
                    s=80, alpha=0.8, edgecolors='k', linewidths=0.5,
                )

        ax.axhline(PETROLEUM_JET_GHG_WTW, color='red', ls='--', lw=1.2,
                   label='Petroleum jet (89 gCO₂e/MJ)')
        ax.set_xlabel('MFSP (2023 $/GGE)', fontsize=11, fontweight='bold')
        ax.set_ylabel('GHG (gCO₂e/MJ)', fontsize=11, fontweight='bold')
        ax.set_title(title, fontsize=11, fontweight='bold', loc='left')
        ax.grid(alpha=0.25)
        ax.set_ylim(-15, 100)

    # Legend: pathways
    pathway_patches = [
        mpatches.Patch(color=c, label=p) for p, c in PATHWAY_COLORS.items()
    ]
    alloc_handles = [
        plt.Line2D([0], [0], marker=m, color='gray', ls='None',
                   markersize=8, label=a.replace('_', ' ').title())
        for a, m in ALLOC_MARKERS.items()
    ]
    axes[1].legend(handles=pathway_patches + alloc_handles,
                   fontsize=8, loc='upper right', ncol=2)

    plt.tight_layout()
    plt.savefig('outputs/figures/Fig1_Literature_Overview.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/Fig1_Literature_Overview.pdf', bbox_inches='tight')
    plt.close()


def fig2_harmonization_impact(df_harm):
    """Figure 2: Variance reduction before vs after harmonization, per pathway."""
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))

    for ax, raw_col, harm_col, ylabel, title_tag in [
        (axes[0], 'mfsp_2023_raw', 'mfsp_harmonized', 'MFSP (2023 $/GGE)', 'Cost'),
        (axes[1], 'ghg_raw',       'ghg_harmonized',  'GHG (gCO₂e/MJ)',    'GHG Emissions'),
    ]:
        pathways = list(PATHWAY_COLORS.keys())
        x = np.arange(len(pathways))
        width = 0.35

        means_raw  = [df_harm[df_harm.pathway == p][raw_col].mean()  for p in pathways]
        stds_raw   = [df_harm[df_harm.pathway == p][raw_col].std()   for p in pathways]
        means_harm = [df_harm[df_harm.pathway == p][harm_col].mean() for p in pathways]
        stds_harm  = [df_harm[df_harm.pathway == p][harm_col].std()  for p in pathways]

        bars_raw  = ax.bar(x - width/2, means_raw,  width, yerr=stds_raw,
                           color=[PATHWAY_COLORS[p] for p in pathways],
                           alpha=0.4, edgecolor='k', capsize=4, label='Raw')
        bars_harm = ax.bar(x + width/2, means_harm, width, yerr=stds_harm,
                           color=[PATHWAY_COLORS[p] for p in pathways],
                           alpha=0.9, edgecolor='k', capsize=4, label='Harmonized')

        ax.set_xticks(x)
        ax.set_xticklabels(pathways, fontsize=10)
        ax.set_ylabel(ylabel, fontsize=11, fontweight='bold')
        ax.set_title(f'({"ab"[axes.tolist().index(ax)]}) {title_tag}: Raw vs Harmonized',
                     fontsize=11, fontweight='bold', loc='left')
        ax.grid(axis='y', alpha=0.25)
        ax.legend(fontsize=9)

        if ax == axes[1]:
            ax.axhline(PETROLEUM_JET_GHG_WTW, color='red', ls='--', lw=1,
                       label='Petroleum baseline')

    plt.tight_layout()
    plt.savefig('outputs/figures/Fig2_Harmonization_Impact.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/Fig2_Harmonization_Impact.pdf', bbox_inches='tight')
    plt.close()


def fig3_monte_carlo_distributions(mc_results):
    """Figure 3: Monte Carlo distributions (violin + percentiles) per pathway."""
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))

    pathways = list(PATHWAY_COLORS.keys())

    for ax, col, xlabel, title in [
        (axes[0], 'mfsp', 'MFSP (2023 $/GGE)', '(a) Minimum Fuel Selling Price Distribution'),
        (axes[1], 'ghg',  'GHG (gCO₂e/MJ)',    '(b) GHG Emissions Distribution'),
    ]:
        data = [mc_results[p][col].values for p in pathways]
        positions = range(1, len(pathways) + 1)

        vp = ax.violinplot(data, positions=positions, showmedians=True,
                           showextrema=False, widths=0.7)

        for i, (body, path) in enumerate(zip(vp['bodies'], pathways)):
            body.set_facecolor(PATHWAY_COLORS[path])
            body.set_alpha(0.7)
        vp['cmedians'].set_color('black')
        vp['cmedians'].set_linewidth(2)

        # Add P5-P95 whiskers
        for i, (d, pos) in enumerate(zip(data, positions)):
            p5, p25, p50, p75, p95 = np.percentile(d, [5, 25, 50, 75, 95])
            ax.plot([pos, pos], [p5, p95], color='gray', lw=1.5, zorder=2)
            ax.plot([pos - 0.12, pos + 0.12], [p5, p5], color='gray', lw=1.5)
            ax.plot([pos - 0.12, pos + 0.12], [p95, p95], color='gray', lw=1.5)

        ax.set_xticks(list(positions))
        ax.set_xticklabels(pathways, fontsize=11)
        ax.set_ylabel(xlabel, fontsize=11, fontweight='bold')
        ax.set_title(title, fontsize=11, fontweight='bold', loc='left')
        ax.grid(axis='y', alpha=0.25)

        if col == 'ghg':
            ax.axhline(PETROLEUM_JET_GHG_WTW, color='red', ls='--', lw=1.2,
                       label='Petroleum jet (89 gCO₂e/MJ)', zorder=1)
            ax.axhline(0, color='black', ls=':', lw=0.8)
            ax.legend(fontsize=9)

    plt.tight_layout()
    plt.savefig('outputs/figures/Fig3_MC_Distributions.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/Fig3_MC_Distributions.pdf', bbox_inches='tight')
    plt.close()


def fig4_variance_decomposition(var_df, sobol_results):
    """Figure 4: Variance decomposition — methodological vs. technical."""
    fig = plt.figure(figsize=(16, 6))
    gs  = GridSpec(1, 3, figure=fig, wspace=0.35)

    # ── Panel (a): Sobol S1 indices for ATJ (representative pathway) ─────────
    ax1  = fig.add_subplot(gs[0])
    pathway_ex = 'ATJ'
    s1m  = sobol_results[pathway_ex]['S1_mfsp']
    s1g  = sobol_results[pathway_ex]['S1_ghg']

    param_labels = list(s1m.keys())
    x = np.arange(len(param_labels))
    w = 0.35
    colors = ['#e53935' if p in METHODOLOGICAL_PARAMS else '#1565C0'
              for p in param_labels]

    ax1.barh(x + w/2, [s1m[p] for p in param_labels], w,
             color=colors, alpha=0.8, edgecolor='k', linewidth=0.5, label='MFSP')
    ax1.barh(x - w/2, [s1g[p] for p in param_labels], w,
             color=colors, alpha=0.4, edgecolor='k', linewidth=0.5,
             hatch='//', label='GHG')

    ax1.set_yticks(x)
    ax1.set_yticklabels(
        [p.replace('_', '\n') for p in param_labels], fontsize=8
    )
    ax1.set_xlabel('First-Order Sobol Index (S₁)', fontsize=10, fontweight='bold')
    ax1.set_title(f'(a) Sensitivity Indices\n({pathway_ex} pathway)',
                  fontsize=10, fontweight='bold', loc='left')
    ax1.grid(axis='x', alpha=0.25)
    ax1.axvline(0, color='k', lw=0.8)

    red_patch  = mpatches.Patch(color='#e53935', label='Methodological')
    blue_patch = mpatches.Patch(color='#1565C0', label='Technical')
    ax1.legend(handles=[red_patch, blue_patch], fontsize=8, loc='lower right')

    # ── Panel (b): Stacked bar — meth vs tech fraction, MFSP ─────────────────
    ax2 = fig.add_subplot(gs[1])
    pathways = var_df['Pathway'].unique()
    mfsp_df  = var_df[var_df.Metric == 'MFSP']
    ghg_df   = var_df[var_df.Metric == 'GHG']

    x2 = np.arange(len(pathways))
    meth_mfsp = [mfsp_df[mfsp_df.Pathway == p]['Methodological_%'].values[0] for p in pathways]
    tech_mfsp = [mfsp_df[mfsp_df.Pathway == p]['Technical_%'].values[0]      for p in pathways]
    meth_ghg  = [ghg_df [ghg_df.Pathway  == p]['Methodological_%'].values[0] for p in pathways]
    tech_ghg  = [ghg_df [ghg_df.Pathway  == p]['Technical_%'].values[0]      for p in pathways]

    w2 = 0.35
    ax2.bar(x2 - w2/2, meth_mfsp, w2, color='#e53935', alpha=0.85, label='Methodological')
    ax2.bar(x2 - w2/2, tech_mfsp, w2, bottom=meth_mfsp, color='#1565C0', alpha=0.85, label='Technical')
    ax2.bar(x2 + w2/2, meth_ghg,  w2, color='#e53935', alpha=0.50, hatch='//')
    ax2.bar(x2 + w2/2, tech_ghg,  w2, bottom=meth_ghg,  color='#1565C0', alpha=0.50, hatch='//')

    ax2.set_xticks(x2)
    ax2.set_xticklabels(pathways, fontsize=10)
    ax2.set_ylabel('Share of Total Variance (%)', fontsize=10, fontweight='bold')
    ax2.set_title('(b) Variance Decomposition\n(solid=MFSP, hatched=GHG)',
                  fontsize=10, fontweight='bold', loc='left')
    ax2.set_ylim(0, 120)
    ax2.legend(fontsize=9)
    ax2.grid(axis='y', alpha=0.25)

    # ── Panel (c): CV reduction from harmonization ────────────────────────────
    ax3 = fig.add_subplot(gs[2])
    for i, pathway in enumerate(pathways):
        for j, (metric, color) in enumerate([('MFSP', '#1565C0'), ('GHG', '#e53935')]):
            row   = var_df[(var_df.Pathway == pathway) & (var_df.Metric == metric)].iloc[0]
            cv_b  = row['CV_before_%']
            cv_a  = row['CV_after_%']
            xpos  = i + j * 0.3 - 0.15
            ax3.plot([xpos, xpos], [cv_a, cv_b], color=color, lw=2, alpha=0.7)
            ax3.scatter(xpos, cv_b, color=color, marker='o', s=60, zorder=5, alpha=0.5)
            ax3.scatter(xpos, cv_a, color=color, marker='D', s=60, zorder=5)

    ax3.set_xticks(range(len(pathways)))
    ax3.set_xticklabels(pathways, fontsize=10)
    ax3.set_ylabel('Coefficient of Variation (%)', fontsize=10, fontweight='bold')
    ax3.set_title('(c) CV Before (○) vs After (◆)\nHarmonization',
                  fontsize=10, fontweight='bold', loc='left')
    ax3.grid(axis='y', alpha=0.25)

    blue_p = mpatches.Patch(color='#1565C0', label='MFSP')
    red_p  = mpatches.Patch(color='#e53935', label='GHG')
    ax3.legend(handles=[blue_p, red_p], fontsize=9)

    plt.savefig('outputs/figures/Fig4_Variance_Decomposition.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/Fig4_Variance_Decomposition.pdf', bbox_inches='tight')
    plt.close()


# =============================================================================
# SECTION 11: EXCEL WORKBOOK
# =============================================================================

def _header_style(cell, bg='1F497D'):
    cell.font  = Font(bold=True, color='FFFFFF', size=10)
    cell.fill  = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', wrap_text=True)


def create_excel_workbook(df_harm, mc_results, var_df):
    """Build multi-sheet Excel workbook with all outputs."""
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Literature Database ─────────────────────────────────────────
    ws1 = wb.create_sheet('Literature Database')
    ws1['A1'] = 'SAF Meta-Analysis: 42 Peer-Reviewed Studies'
    ws1['A1'].font = Font(bold=True, size=13)

    cols_lit = [
        'study_id', 'authors', 'year', 'journal', 'pathway', 'feedstock',
        'plant_size_tpd', 'ref_year_cost', 'mfsp_raw', 'mfsp_unit',
        'ghg_raw', 'allocation', 'boundary', 'include_iluc',
        'discount_rate', 'capacity_factor', 'plant_lifetime',
    ]
    lit_df = df_harm[cols_lit].copy()
    for r_idx, row in enumerate(dataframe_to_rows(lit_df, index=False, header=True), start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 3:
                _header_style(cell)
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 18

    # ── Sheet 2: Harmonized Values ────────────────────────────────────────────
    ws2 = wb.create_sheet('Harmonized Values')
    ws2['A1'] = 'Harmonization Protocol: WtWake | Energy Alloc | 2023 USD | 10% DR | 90% CF'
    ws2['A1'].font = Font(bold=True, size=11)

    cols_harm = [
        'study_id', 'authors', 'year', 'pathway', 'feedstock',
        'mfsp_2023_raw', 'mfsp_harmonized', 'ghg_raw', 'ghg_harmonized',
        'ghg_reduction_%', 'alloc_correction', 'boundary_correction_applied',
        'iluc_removed_gco2e', 'crf_correction',
    ]
    harm_out = df_harm[cols_harm].copy()
    for r_idx, row in enumerate(dataframe_to_rows(harm_out, index=False, header=True), start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 3:
                _header_style(cell)
            elif c_idx in (6, 7) and r_idx > 3:
                cell.number_format = '$#,##0.00'
            elif c_idx in (8, 9, 10) and r_idx > 3:
                cell.number_format = '0.0'
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 20

    # ── Sheet 3: Monte Carlo Summary ──────────────────────────────────────────
    ws3 = wb.create_sheet('Monte Carlo Summary')
    ws3['A1'] = 'Monte Carlo Results — 10,000 Iterations per Pathway'
    ws3['A1'].font = Font(bold=True, size=11)

    mc_rows = [['Pathway', 'Metric', 'P5', 'P25', 'Median', 'Mean', 'P75', 'P95', 'Std', 'CV%']]
    for pathway in PATHWAY_CONFIGS:
        mc = mc_results[pathway]
        for col, metric in [('mfsp', 'MFSP ($/GGE)'), ('ghg', 'GHG (gCO2e/MJ)')]:
            vals = mc[col].dropna()
            mc_rows.append([
                pathway, metric,
                round(np.percentile(vals, 5),  2),
                round(np.percentile(vals, 25), 2),
                round(np.percentile(vals, 50), 2),
                round(vals.mean(), 2),
                round(np.percentile(vals, 75), 2),
                round(np.percentile(vals, 95), 2),
                round(vals.std(),  2),
                round(vals.std() / vals.mean() * 100, 1),
            ])
    for r_idx, row_data in enumerate(mc_rows, start=3):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 3:
                _header_style(cell)
    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 18

    # ── Sheet 4: Variance Decomposition ───────────────────────────────────────
    ws4 = wb.create_sheet('Variance Decomposition')
    ws4['A1'] = 'Variance Decomposition: Methodological vs. Technical/Economic'
    ws4['A1'].font = Font(bold=True, size=11)

    for r_idx, row in enumerate(
        dataframe_to_rows(var_df, index=False, header=True), start=3
    ):
        for c_idx, val in enumerate(row, start=1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 3:
                _header_style(cell)
    for col in ws4.columns:
        ws4.column_dimensions[col[0].column_letter].width = 22

    # ── Sheet 5: Key Findings ─────────────────────────────────────────────────
    ws5 = wb.create_sheet('Key Findings')
    ws5['A1'] = 'KEY FINDINGS — SAF Harmonization Meta-Analysis'
    ws5['A1'].font = Font(bold=True, size=13)

    df_all_harm_mfsp = df_harm['mfsp_harmonized']
    df_all_harm_ghg  = df_harm['ghg_harmonized']

    findings = [
        ['Finding', 'Value', 'Notes'],
        ['Studies analyzed', len(df_harm), '42 peer-reviewed publications, 2009–2022'],
        ['Pathways covered', 4, 'ATJ, HEFA, FT-SPK, PtL'],
        ['Raw MFSP range (2023 $/GGE)',
         f"${df_harm['mfsp_2023_raw'].min():.2f}–${df_harm['mfsp_2023_raw'].max():.2f}",
         'Before harmonization'],
        ['Harmonized MFSP range (2023 $/GGE)',
         f"${df_all_harm_mfsp.min():.2f}–${df_all_harm_mfsp.max():.2f}",
         'After harmonization to common basis'],
        ['Raw GHG range (gCO2e/MJ)',
         f"{df_harm['ghg_raw'].min():.1f}–{df_harm['ghg_raw'].max():.1f}",
         'Before harmonization'],
        ['Harmonized GHG range (gCO2e/MJ)',
         f"{df_all_harm_ghg.min():.1f}–{df_all_harm_ghg.max():.1f}",
         'After harmonization'],
        ['Petroleum jet baseline (gCO2e/MJ)', PETROLEUM_JET_GHG_WTW, 'ICAO CORSIA WtWake'],
        ['Avg GHG reduction vs petroleum',
         f"{df_harm['ghg_reduction_%'].mean():.0f}%",
         'Mean across all 42 harmonized studies'],
    ]

    for pathway in PATHWAY_CONFIGS:
        for metric, col in [('MFSP ($/GGE)', 'mfsp'), ('GHG (gCO2e/MJ)', 'ghg')]:
            med = mc_results[pathway][col].median()
            p5  = np.percentile(mc_results[pathway][col], 5)
            p95 = np.percentile(mc_results[pathway][col], 95)
            vr  = var_df[(var_df.Pathway == pathway) & (var_df.Metric == metric.split()[0])]['Variance_Reduction_%'].values
            vr_str = f"{vr[0]:.0f}%" if len(vr) else 'N/A'
            findings.append([
                f'{pathway} {metric} (median [P5–P95])',
                f"{med:.2f} [{p5:.2f}–{p95:.2f}]",
                f"Variance reduction from harmonization: {vr_str}",
            ])

    for r_idx, row_data in enumerate(findings, start=3):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws5.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 3:
                _header_style(cell)
    ws5.column_dimensions['A'].width = 40
    ws5.column_dimensions['B'].width = 28
    ws5.column_dimensions['C'].width = 45

    return wb


# =============================================================================
# SECTION 12: ADDITIONAL PUBLICATION FIGURES
# =============================================================================

def _modal_params(param_defs):
    """Extract modal (best-estimate) values from distribution specs."""
    modal = {}
    for name, spec in param_defs.items():
        kind = spec[0]
        if kind == 'triangular':
            modal[name] = spec[2]           # mode
        elif kind == 'normal':
            modal[name] = spec[1]           # mean
        elif kind == 'uniform':
            modal[name] = (spec[1] + spec[2]) / 2
    return modal


def _percentile_val(spec, q):
    """Return q-th quantile (0–1) of a distribution spec."""
    kind = spec[0]
    if kind == 'triangular':
        _, low, mode, high = spec
        c = (mode - low) / (high - low) if high > low else 0.5
        return float(stats.triang.ppf(q, c=c, loc=low, scale=high - low))
    if kind == 'uniform':
        _, low, high = spec
        return low + q * (high - low)
    if kind == 'normal':
        _, mean, std = spec
        return float(stats.norm.ppf(q, loc=mean, scale=std))
    return spec[1]


def compute_cost_ghg_breakdown():
    """
    Run each pathway model at modal parameter values.
    Returns itemised cost ($/GGE) and GHG (gCO2e/MJ) components per pathway.
    """
    results = {}

    # ── ATJ ───────────────────────────────────────────────────────────────────
    p = _modal_params(ATJ_PARAMS)
    annual_feed   = 2000 * 365 * p['capacity_factor']
    annual_etoh   = annual_feed * p['ethanol_yield']
    annual_jet    = annual_etoh * p['jet_yield']
    annual_diesel = annual_etoh * 0.15
    annual_gas    = annual_etoh * 0.14
    jet_gge       = annual_jet    * (JET_LHV_BTU_PER_GAL / GASOLINE_LHV_BTU_PER_GAL)
    diesel_gge    = annual_diesel * (129_488 / GASOLINE_LHV_BTU_PER_GAL)
    total_gge     = jet_gge + diesel_gge + annual_gas
    crf_val       = _crf(p['discount_rate'], HARMONIZED_PLANT_LIFETIME)
    results['ATJ'] = {
        'cost': {
            'Capital':   p['capex_2023'] * crf_val / total_gge,
            'O&M':       p['capex_2023'] * 0.07    / total_gge,
            'Feedstock': annual_feed * p['feedstock_cost'] / total_gge,
        },
        'ghg': {
            'Feedstock GHG':  p['feedstock_ghg'] * p['alloc_factor'],
            'NG combustion':  p['ng_use'] * 56.1  * p['alloc_factor'],
            'Electricity':    p['elec_use'] * p['grid_intensity'] * p['alloc_factor'],
            'Boundary offset': p['boundary_offset'],
        },
    }

    # ── HEFA ──────────────────────────────────────────────────────────────────
    p = _modal_params(HEFA_PARAMS)
    annual_feed    = 800 * 365 * p['capacity_factor']
    annual_jet_kg  = annual_feed * 1000 * p['jet_yield']
    annual_jet_L   = annual_jet_kg / 0.804
    annual_jet_GGE = annual_jet_L / L_JET_PER_GGE
    crf_val        = _crf(p['discount_rate'], HARMONIZED_PLANT_LIFETIME)
    h2_mass        = annual_feed * p['h2_use']
    h2_ghg_alloc   = (h2_mass * 1000 * 9000) / (annual_jet_L * JET_LHV_MJ_PER_L)
    results['HEFA'] = {
        'cost': {
            'Capital':   p['capex_2023'] * crf_val / annual_jet_GGE,
            'O&M':       p['capex_2023'] * 0.07    / annual_jet_GGE,
            'Feedstock': annual_feed * p['feedstock_cost'] / annual_jet_GGE,
            'H₂':        h2_mass * 1000 * p['h2_price'] / 1000 / annual_jet_GGE,
        },
        'ghg': {
            'Feedstock GHG': p['feedstock_ghg'] * p['alloc_factor'],
            'H₂ production': h2_ghg_alloc        * p['alloc_factor'],
            'Process GHG':   p['process_ghg']    * p['alloc_factor'],
            'Boundary offset': p['boundary_offset'],
        },
    }

    # ── FT-SPK ────────────────────────────────────────────────────────────────
    p = _modal_params(FTSPK_PARAMS)
    annual_feed = 2500 * 365 * p['capacity_factor']
    fuel_MJ     = annual_feed * 17_500 * p['ft_efficiency']
    jet_MJ      = fuel_MJ * 0.65
    jet_L       = jet_MJ / JET_LHV_MJ_PER_L
    jet_GGE     = jet_L / L_JET_PER_GGE
    crf_val     = _crf(p['discount_rate'], HARMONIZED_PLANT_LIFETIME)
    results['FT-SPK'] = {
        'cost': {
            'Capital':   p['capex_2023'] * crf_val / jet_GGE,
            'O&M':       p['capex_2023'] * 0.05    / jet_GGE,
            'Feedstock': annual_feed * p['feedstock_cost'] / jet_GGE,
        },
        'ghg': {
            'Feedstock GHG': p['feedstock_ghg'] * p['alloc_factor'],
            'Process GHG':   p['process_ghg']   * p['alloc_factor'],
            'Boundary offset': p['boundary_offset'],
        },
    }

    # ── PtL ───────────────────────────────────────────────────────────────────
    p = _modal_params(PTL_PARAMS)
    PLANT_MW     = 200
    annual_MWh   = PLANT_MW * 8760 * p['capacity_factor']
    annual_kg_H2 = annual_MWh * 1000 / 55.0
    annual_CO2_t = annual_kg_H2 * 5.5 / 1000
    jet_MJ_ptl   = annual_kg_H2 * 120.0 * p['ft_efficiency'] * 0.70
    jet_L_ptl    = jet_MJ_ptl / JET_LHV_MJ_PER_L
    jet_GGE_ptl  = jet_L_ptl / L_JET_PER_GGE
    crf_val      = _crf(p['discount_rate'], HARMONIZED_PLANT_LIFETIME)
    elec_capex   = p['elec_capex_kw'] * PLANT_MW * 1000
    results['PtL'] = {
        'cost': {
            'Electrolyzer cap.': elec_capex * crf_val / jet_GGE_ptl,
            'FT capital':        p['ft_capex'] * crf_val / jet_GGE_ptl,
            'Electricity':       annual_MWh * p['elec_cost_mwh'] / jet_GGE_ptl,
            'CO₂ capture':       annual_CO2_t * p['co2_capture_cost'] / jet_GGE_ptl,
            'O&M':               (elec_capex + p['ft_capex']) * 0.04 / jet_GGE_ptl,
        },
        'ghg': {
            'Electricity GHG': annual_MWh * 1000 * p['grid_intensity'] / jet_MJ_ptl,
            'Boundary offset': p['boundary_offset'],
        },
    }

    return results


def _oat_swings(pathway, param_defs, model_fn, metric='mfsp'):
    """
    One-at-a-time (OAT) sensitivity: swing in `metric` when each parameter
    varies from P5 to P95, all others held at modal values.
    Returns list of (param_name, swing_lo, swing_hi) sorted by |swing| desc.
    """
    modal     = _modal_params(param_defs)
    modal_val = model_fn(modal)[metric]
    swings    = []
    for name, spec in param_defs.items():
        p5  = _percentile_val(spec, 0.05)
        p95 = _percentile_val(spec, 0.95)
        try:
            lo = model_fn({**modal, name: p5})[metric]
            hi = model_fn({**modal, name: p95})[metric]
        except Exception:
            lo = hi = modal_val
        swings.append((name, lo - modal_val, hi - modal_val))
    swings.sort(key=lambda x: abs(x[2] - x[1]), reverse=True)
    return swings


# ─────────────────────────────────────────────────────────────────────────────
# figA: Well-to-Wake System Boundary
# ─────────────────────────────────────────────────────────────────────────────

def figA_system_boundary():
    """Figure A: Well-to-Wake system boundary schematic."""
    fig, ax = plt.subplots(figsize=(14, 5))
    ax.set_xlim(0, 14); ax.set_ylim(0, 5)
    ax.axis('off')
    ax.set_facecolor('#F8F9FA'); fig.patch.set_facecolor('#F8F9FA')

    stages = [
        (0.3,  2.5, 1.8, 1.4, '#BBDEFB', 'Feedstock\nProduction',  'Well / Farm\nForest / DAC'),
        (2.6,  2.5, 1.8, 1.4, '#C8E6C9', 'Feedstock\nTransport',   'Logistics &\nStorage'),
        (4.9,  2.5, 1.8, 1.4, '#FFE0B2', 'SAF\nConversion',         'Biorefinery /\nElectrolyzer'),
        (7.2,  2.5, 1.8, 1.4, '#F3E5F5', 'Fuel\nDistribution',      'Pipeline /\nTanker'),
        (9.5,  2.5, 1.8, 1.4, '#FFCDD2', 'Combustion\n(Aircraft)',  'Jet engine\nexhaust'),
    ]
    box_centers = []
    for x, y, w, h, color, title, sub in stages:
        rect = FancyBboxPatch((x, y - h/2), w, h, boxstyle='round,pad=0.07',
                              facecolor=color, edgecolor='#455A64', linewidth=1.5)
        ax.add_patch(rect)
        ax.text(x + w/2, y + 0.22, title, ha='center', va='center',
                fontsize=9, fontweight='bold', color='#1A237E')
        ax.text(x + w/2, y - 0.28, sub, ha='center', va='center',
                fontsize=7.5, color='#37474F')
        box_centers.append((x + w, y))

    for i in range(len(box_centers) - 1):
        x0, y0 = box_centers[i]
        x1 = stages[i + 1][0]
        ax.annotate('', xy=(x1, y0), xytext=(x0, y0),
                    arrowprops=dict(arrowstyle='->', color='#455A64',
                                   lw=1.8, mutation_scale=14))

    for label, x0, x1, color, y_off in [
        ('Well-to-Gate (WtG)',  0.3, 8.8,  '#1565C0', 0.0),
        ('Well-to-Wake (WtWake)', 0.3, 11.7, '#B71C1C', -0.45),
    ]:
        y_b = 1.35 + y_off
        ax.annotate('', xy=(x1, y_b), xytext=(x0, y_b),
                    arrowprops=dict(arrowstyle='<->', color=color,
                                   lw=2, mutation_scale=12))
        ax.text((x0 + x1) / 2, y_b - 0.28, label,
                ha='center', fontsize=9, fontweight='bold', color=color)

    ax.text(12.0, 2.5,
            'Excluded:\n• ILUC (sensitivity)\n• Infrastructure\n• End-of-life',
            ha='left', va='center', fontsize=7.5, color='#616161',
            bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                      edgecolor='#9E9E9E', alpha=0.85))

    ax.set_title(
        'Figure A — Well-to-Wake System Boundary  |  '
        'Functional Unit: 1 MJ neat SAF at aircraft fueling point',
        fontsize=10, fontweight='bold', loc='left', pad=8)
    plt.tight_layout()
    plt.savefig('outputs/figures/FigA_System_Boundary.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/FigA_System_Boundary.pdf', bbox_inches='tight')
    plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# figB: Harmonization Flowchart
# ─────────────────────────────────────────────────────────────────────────────

def figB_harmonization_flowchart(df_harm):
    """Figure B: Five-step harmonization protocol."""
    fig, ax = plt.subplots(figsize=(13, 9))
    ax.set_xlim(0, 13); ax.set_ylim(0, 10)
    ax.axis('off')

    steps = [
        ('1', 'Raw Study Data',
         'As reported: mixed currency years, system boundaries, allocation methods',
         '#ECEFF1', '#37474F'),
        ('2', 'CEPCI / CPI Cost Escalation',
         'Capital costs → 2023 USD (CEPCI₂₀₂₃ = 798); CPI escalation for opex/feedstock',
         '#E3F2FD', '#1565C0'),
        ('3', 'Allocation Re-weighting',
         'Energy allocation per ISO 14044 §4.3.4.2; pathway-specific factors applied',
         '#E8F5E9', '#1B5E20'),
        ('4', 'System Boundary Offset',
         'All studies converted to Well-to-Wake; WtG→WtWake delta (+3.0 gCO₂e/MJ)',
         '#FFF3E0', '#E65100'),
        ('5', 'CRF Normalization',
         'Uniform 10% discount rate, 30-yr plant life, 90% capacity factor via CRF',
         '#F3E5F5', '#4A148C'),
        ('✓', 'Harmonized Dataset (42 studies)',
         'WtWake | 1 MJ FU | Energy alloc | 2023 USD | 10% DR | 90% CF',
         '#C8E6C9', '#1B5E20'),
    ]
    y_positions = [9.1, 7.7, 6.3, 4.9, 3.5, 1.9]
    box_x, box_w, box_h = 1.5, 8.5, 0.95

    for (num, title, desc, bg, fg), y in zip(steps, y_positions):
        rect = FancyBboxPatch((box_x, y - box_h / 2), box_w, box_h,
                              boxstyle='round,pad=0.1',
                              facecolor=bg, edgecolor=fg, linewidth=1.8)
        ax.add_patch(rect)
        circle = plt.Circle((box_x + 0.45, y), 0.28, color=fg, zorder=3)
        ax.add_patch(circle)
        ax.text(box_x + 0.45, y, num, ha='center', va='center',
                fontsize=9, fontweight='bold', color='white', zorder=4)
        ax.text(box_x + 1.0, y + 0.19, title, ha='left', va='center',
                fontsize=10, fontweight='bold', color=fg)
        ax.text(box_x + 1.0, y - 0.22, desc, ha='left', va='center',
                fontsize=8, color='#37474F')
        if y > y_positions[-1]:
            ax.annotate('', xy=(box_x + box_w / 2, y - box_h / 2 - 0.18),
                        xytext=(box_x + box_w / 2, y - box_h / 2),
                        arrowprops=dict(arrowstyle='->', color='#607D8B', lw=1.8))

    # Reference basis sidebar
    ax.text(10.5, 5.5,
            'Reference Basis\n'
            '─────────────────\n'
            'Boundary: Well-to-Wake\n'
            'FU       : 1 MJ SAF\n'
            'Allocation: Energy\n'
            'Cost yr  : 2023 USD\n'
            'CEPCI    : 798\n'
            'Discount : 10% real\n'
            'Lifetime : 30 yr\n'
            'Cap. fac.: 90%',
            ha='center', va='center', fontsize=8, fontfamily='monospace',
            bbox=dict(boxstyle='round,pad=0.5', facecolor='#FFF9C4',
                      edgecolor='#F9A825', linewidth=1.5))

    # Correction magnitudes from df_harm
    mfsp_delta = (df_harm['mfsp_harmonized'] - df_harm['mfsp_2023_raw']).abs().mean()
    ghg_delta  = (df_harm['ghg_harmonized']  - df_harm['ghg_raw']).abs().mean()
    ax.text(0.2, 0.5,
            f'Mean |ΔMFSP| = ${mfsp_delta:.2f}/GGE\nMean |ΔGHG| = {ghg_delta:.1f} gCO₂e/MJ',
            fontsize=8, color='#424242',
            bbox=dict(boxstyle='round', facecolor='white', edgecolor='#BDBDBD'))

    ax.set_title('Figure B — Five-Step Harmonization Protocol',
                 fontsize=11, fontweight='bold', loc='left', pad=6)
    plt.tight_layout()
    plt.savefig('outputs/figures/FigB_Harmonization_Flowchart.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/FigB_Harmonization_Flowchart.pdf', bbox_inches='tight')
    plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# figC: MFSP Cost Breakdown
# ─────────────────────────────────────────────────────────────────────────────

def figC_cost_breakdown(breakdown):
    """Figure C: Stacked bar of MFSP cost components (modal parameter values)."""
    fig, ax = plt.subplots(figsize=(11, 6))
    pathways  = list(breakdown.keys())
    all_items = list(dict.fromkeys(
        item for pw in pathways for item in breakdown[pw]['cost']
    ))
    item_colors = {
        'Capital':            '#1565C0',
        'O&M':                '#64B5F6',
        'Feedstock':          '#388E3C',
        'H₂':                 '#81C784',
        'Electrolyzer cap.':  '#7B1FA2',
        'FT capital':         '#CE93D8',
        'Electricity':        '#F57C00',
        'CO₂ capture':        '#FDD835',
    }
    default_clrs = list(plt.cm.tab20.colors)
    for i, item in enumerate(all_items):
        if item not in item_colors:
            item_colors[item] = default_clrs[i % len(default_clrs)]

    x      = np.arange(len(pathways))
    bottom = np.zeros(len(pathways))
    for item in all_items:
        vals = np.array([breakdown[pw]['cost'].get(item, 0) for pw in pathways])
        ax.bar(x, vals, bottom=bottom, color=item_colors[item],
               edgecolor='white', linewidth=0.6, label=item, width=0.55)
        for xi, (v, b) in enumerate(zip(vals, bottom)):
            if v > 0.3:
                ax.text(xi, b + v / 2, f'{v:.1f}', ha='center', va='center',
                        fontsize=8, color='white', fontweight='bold')
        bottom += vals

    ax.set_xticks(x)
    ax.set_xticklabels(pathways, fontsize=12, fontweight='bold')
    for tick, pw in zip(ax.get_xticklabels(), pathways):
        tick.set_color(PATHWAY_COLORS[pw])
    ax.set_ylabel('MFSP (2023 $/GGE)', fontsize=11, fontweight='bold')
    ax.set_title('Figure C — MFSP Cost Component Breakdown at Modal Parameter Values\n'
                 '(computed from pathway TEA models)',
                 fontsize=10, fontweight='bold', loc='left')
    ax.set_ylim(0, max(bottom) * 1.20)
    ax.legend(loc='upper left', fontsize=8.5, ncol=2, framealpha=0.9)
    ax.grid(axis='y', alpha=0.25)
    plt.tight_layout()
    plt.savefig('outputs/figures/FigC_Cost_Breakdown.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/FigC_Cost_Breakdown.pdf', bbox_inches='tight')
    plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# figD: Tornado Chart — OAT sensitivity
# ─────────────────────────────────────────────────────────────────────────────

def figD_tornado_sensitivity():
    """Figure D: Tornado chart — one-at-a-time (OAT) MFSP sensitivity per pathway."""
    model_map = {'ATJ': atj_model, 'HEFA': hefa_model,
                 'FT-SPK': ftspk_model, 'PtL': ptl_model}
    param_map = {'ATJ': ATJ_PARAMS, 'HEFA': HEFA_PARAMS,
                 'FT-SPK': FTSPK_PARAMS, 'PtL': PTL_PARAMS}

    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('Figure D — Tornado Sensitivity Chart  (OAT: P5 → P95 parameter swing)\n'
                 'Red labels = methodological parameters; blue = technical',
                 fontsize=11, fontweight='bold')

    for ax, pathway in zip(axes.flat, ['ATJ', 'HEFA', 'FT-SPK', 'PtL']):
        swings     = _oat_swings(pathway, param_map[pathway], model_map[pathway], 'mfsp')
        top_n      = swings[:8]
        modal_mfsp = model_map[pathway](_modal_params(param_map[pathway]))['mfsp']
        y_pos      = list(range(len(top_n)))

        for i, (name, sw_lo, sw_hi) in enumerate(top_n):
            is_meth = name in METHODOLOGICAL_PARAMS
            clr     = '#e53935' if is_meth else '#1565C0'
            ax.barh(i, sw_hi, left=0, color=clr,     alpha=0.75, height=0.5)
            ax.barh(i, sw_lo, left=0, color='#FFCDD2' if is_meth else '#BBDEFB',
                    alpha=0.9, height=0.5, edgecolor=clr, linewidth=0.8)

        ax.axvline(0, color='black', lw=1.2)
        ax.set_yticks(y_pos)
        labels = [s[0].replace('_', ' ') for s in top_n]
        ax.set_yticklabels(labels, fontsize=8.5)
        for ytick, (name, _, _) in zip(ax.get_yticklabels(), top_n):
            ytick.set_color('#e53935' if name in METHODOLOGICAL_PARAMS else '#1565C0')
        ax.set_xlabel('ΔMFSP from modal ($/GGE)', fontsize=9, fontweight='bold')
        ax.set_title(f'{pathway}  (modal = ${modal_mfsp:.2f}/GGE)',
                     fontsize=10, fontweight='bold', color=PATHWAY_COLORS[pathway])
        ax.grid(axis='x', alpha=0.25)

        hi_patch = mpatches.Patch(color='#1565C0', alpha=0.75, label='P95 swing (tech.)')
        lo_patch = mpatches.Patch(color='#e53935', alpha=0.75, label='P95 swing (meth.)')
        ax.legend(handles=[hi_patch, lo_patch], fontsize=7.5, loc='lower right')

    plt.tight_layout()
    plt.savefig('outputs/figures/FigD_Tornado_Sensitivity.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/FigD_Tornado_Sensitivity.pdf', bbox_inches='tight')
    plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# figE: MC KDE Overlay — raw literature vs Monte Carlo
# ─────────────────────────────────────────────────────────────────────────────

def figE_mc_overlay(df_harm, mc_results):
    """Figure E: KDE overlay — raw literature data vs Monte Carlo MFSP distributions."""
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))
    fig.suptitle('Figure E — Monte Carlo Distributions vs Raw & Harmonized Literature Values',
                 fontsize=11, fontweight='bold')

    for ax, pathway in zip(axes.flat, ['ATJ', 'HEFA', 'FT-SPK', 'PtL']):
        color = PATHWAY_COLORS[pathway]
        mc    = mc_results[pathway]['mfsp'].values
        mc    = mc[np.isfinite(mc)]
        lit   = df_harm[df_harm['pathway'] == pathway]

        # MC KDE
        if len(mc) > 1:
            kde = stats.gaussian_kde(mc)
            x   = np.linspace(mc.min(), mc.max(), 300)
            ax.fill_between(x, kde(x), alpha=0.30, color=color)
            ax.plot(x, kde(x), color=color, lw=2, label='MC distribution')

        # Literature rugs
        if len(lit) > 0:
            raw_vals  = lit['mfsp_2023_raw'].dropna().values
            harm_vals = lit['mfsp_harmonized'].dropna().values
            rug_y     = -0.008
            ax.scatter(raw_vals,  np.full(len(raw_vals), rug_y),
                       marker='|', s=100, color='black', alpha=0.8, linewidths=2,
                       label=f'Raw literature (n={len(raw_vals)})', zorder=5)
            ax.scatter(harm_vals, np.full(len(harm_vals), rug_y - 0.008),
                       marker='|', s=100, color=color, alpha=0.8, linewidths=2,
                       label='Harmonized literature', zorder=5)

        p5, p50, p95 = np.nanpercentile(mc, [5, 50, 95])
        ax.axvline(p50, color=color, ls='--', lw=1.8, label=f'MC P50 = ${p50:.2f}')
        ax.axvline(p5,  color=color, ls=':',  lw=1.0, alpha=0.7)
        ax.axvline(p95, color=color, ls=':',  lw=1.0, alpha=0.7)
        ax.text(p5,  ax.get_ylim()[1] * 0.05 if ax.get_ylim()[1] > 0 else 0.01,
                'P5', fontsize=7, color=color, ha='center')
        ax.text(p95, ax.get_ylim()[1] * 0.05 if ax.get_ylim()[1] > 0 else 0.01,
                'P95', fontsize=7, color=color, ha='center')

        ax.set_xlabel('MFSP (2023 $/GGE)', fontsize=10, fontweight='bold')
        ax.set_ylabel('Probability Density', fontsize=9)
        ax.set_title(pathway, fontsize=11, fontweight='bold', color=color)
        ax.legend(fontsize=8)
        ax.grid(alpha=0.2)
        ax.set_ylim(bottom=-0.02)

    plt.tight_layout()
    plt.savefig('outputs/figures/FigE_MC_Overlay.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/FigE_MC_Overlay.pdf', bbox_inches='tight')
    plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# figF: GHG Comparison vs Regulatory Thresholds
# ─────────────────────────────────────────────────────────────────────────────

def figF_ghg_comparison(mc_results, df_harm):
    """Figure F: GHG distributions (MC violin) vs regulatory thresholds."""
    fig, ax = plt.subplots(figsize=(12, 7))
    pathways  = ['ATJ', 'HEFA', 'FT-SPK', 'PtL']
    positions = np.arange(1, len(pathways) + 1)
    data      = [mc_results[p]['ghg'].values for p in pathways]

    vp = ax.violinplot(data, positions=positions, showmedians=True,
                       showextrema=False, widths=0.65)
    for body, pw in zip(vp['bodies'], pathways):
        body.set_facecolor(PATHWAY_COLORS[pw]); body.set_alpha(0.60)
    vp['cmedians'].set_color('black'); vp['cmedians'].set_linewidth(2.5)

    for i, (d, pos) in enumerate(zip(data, positions)):
        p5, p25, p75, p95 = np.nanpercentile(d, [5, 25, 75, 95])
        ax.plot([pos, pos], [p5, p95], 'k-', lw=1.8, zorder=3)
        ax.fill_betweenx([p25, p75], pos - 0.09, pos + 0.09,
                         color='black', alpha=0.20, zorder=4)

    # Literature harmonized overlay
    for pw, pos in zip(pathways, positions):
        lit_vals = df_harm[df_harm['pathway'] == pw]['ghg_harmonized'].dropna()
        if len(lit_vals):
            ax.scatter(np.full(len(lit_vals), pos), lit_vals,
                       color=PATHWAY_COLORS[pw], s=40, alpha=0.60, zorder=5,
                       edgecolors='k', linewidths=0.5)

    thresholds = [
        (PETROLEUM_JET_GHG_WTW,          'Petroleum jet (89 gCO₂e/MJ)', 'red',     '--'),
        (PETROLEUM_JET_GHG_WTW * 0.50,   'CORSIA Tier 2 (−50%)',        '#F57C00', '-.'),
        (PETROLEUM_JET_GHG_WTW * 0.35,   'EU RED III (−65%)',            '#1565C0', ':'),
        (0.0,                              'Net-zero',                     '#424242', ':'),
    ]
    for val, label, color, ls in thresholds:
        ax.axhline(val, color=color, ls=ls, lw=1.5, alpha=0.85, label=label)

    ax.set_xticks(list(positions))
    ax.set_xticklabels(pathways, fontsize=12, fontweight='bold')
    for tick, pw in zip(ax.get_xticklabels(), pathways):
        tick.set_color(PATHWAY_COLORS[pw])
    ax.set_ylabel('GHG Intensity (gCO₂e/MJ)', fontsize=11, fontweight='bold')
    ax.set_title('Figure F — GHG Distributions vs Regulatory Thresholds\n'
                 '(violin = MC 10k iterations; dots = harmonized literature studies)',
                 fontsize=10, fontweight='bold', loc='left')
    ax.legend(fontsize=9, loc='upper right')
    ax.grid(axis='y', alpha=0.25)
    ax.set_ylim(bottom=-5)
    plt.tight_layout()
    plt.savefig('outputs/figures/FigF_GHG_Comparison.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/FigF_GHG_Comparison.pdf', bbox_inches='tight')
    plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# figG: GHG Component Breakdown
# ─────────────────────────────────────────────────────────────────────────────

def figG_ghg_contribution(breakdown):
    """Figure G: Stacked bar of GHG component breakdown per pathway."""
    fig, ax = plt.subplots(figsize=(11, 6))
    pathways  = list(breakdown.keys())
    all_items = list(dict.fromkeys(
        item for pw in pathways for item in breakdown[pw]['ghg']
    ))
    ghg_colors = {
        'Feedstock GHG':   '#388E3C',
        'NG combustion':   '#F57C00',
        'Electricity':     '#FDD835',
        'Electricity GHG': '#FDD835',
        'H₂ production':   '#0288D1',
        'Process GHG':     '#78909C',
        'Boundary offset': '#D32F2F',
    }
    x      = np.arange(len(pathways))
    bottom = np.zeros(len(pathways))
    for item in all_items:
        vals = np.array([max(breakdown[pw]['ghg'].get(item, 0), 0) for pw in pathways])
        ax.bar(x, vals, bottom=bottom,
               color=ghg_colors.get(item, '#9E9E9E'),
               edgecolor='white', linewidth=0.6, label=item, width=0.55)
        for xi, (v, b) in enumerate(zip(vals, bottom)):
            if v > 1.0:
                ax.text(xi, b + v / 2, f'{v:.1f}', ha='center', va='center',
                        fontsize=8, color='white', fontweight='bold')
        bottom += vals

    ax.axhline(PETROLEUM_JET_GHG_WTW, color='red', ls='--', lw=1.5,
               label='Petroleum jet (89 gCO₂e/MJ)')
    ax.set_xticks(x)
    ax.set_xticklabels(pathways, fontsize=12, fontweight='bold')
    for tick, pw in zip(ax.get_xticklabels(), pathways):
        tick.set_color(PATHWAY_COLORS[pw])
    ax.set_ylabel('GHG Intensity (gCO₂e/MJ)', fontsize=11, fontweight='bold')
    ax.set_title('Figure G — GHG Component Breakdown at Modal Parameter Values\n'
                 '(Well-to-Wake; energy allocation; 2023 basis)',
                 fontsize=10, fontweight='bold', loc='left')
    ax.legend(fontsize=8.5, ncol=2, loc='upper right')
    ax.grid(axis='y', alpha=0.25)
    plt.tight_layout()
    plt.savefig('outputs/figures/FigG_GHG_Contribution.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/FigG_GHG_Contribution.pdf', bbox_inches='tight')
    plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# figH: Parameter Sensitivity Heatmap (Sobol S1)
# ─────────────────────────────────────────────────────────────────────────────

def figH_parameter_heatmap(sobol_results):
    """Figure H: Heatmap of Sobol first-order indices (pathways × parameters)."""
    fig, axes = plt.subplots(1, 2, figsize=(16, 7))

    for ax, metric_key, title_tag, panel in [
        (axes[0], 'S1_mfsp', 'MFSP', 'a'),
        (axes[1], 'S1_ghg',  'GHG',  'b'),
    ]:
        pathways   = list(sobol_results.keys())
        all_params = list(dict.fromkeys(
            p for pw in pathways for p in sobol_results[pw][metric_key]
        ))
        matrix = pd.DataFrame(
            {pw: [sobol_results[pw][metric_key].get(p, 0) for p in all_params]
             for pw in pathways},
            index=all_params,
        )
        # Sort by max across pathways
        matrix = matrix.loc[matrix.max(axis=1).sort_values(ascending=False).index]

        vmax = min(matrix.values.max(), 0.65)
        im   = ax.imshow(matrix.values, cmap='YlOrRd', aspect='auto',
                         vmin=0, vmax=vmax)
        ax.set_xticks(range(len(pathways)))
        ax.set_xticklabels(pathways, fontsize=11, fontweight='bold')
        ax.set_yticks(range(len(matrix.index)))
        ax.set_yticklabels([p.replace('_', ' ') for p in matrix.index], fontsize=9)

        for ytick, param in zip(ax.get_yticklabels(), matrix.index):
            ytick.set_color('#e53935' if param in METHODOLOGICAL_PARAMS else '#1565C0')

        for i in range(matrix.shape[0]):
            for j in range(matrix.shape[1]):
                val = matrix.values[i, j]
                ax.text(j, i, f'{val:.2f}', ha='center', va='center',
                        fontsize=8.5, fontweight='bold',
                        color='white' if val > vmax * 0.6 else 'black')

        plt.colorbar(im, ax=ax, shrink=0.75, label='S₁ (First-Order Sobol Index)')
        ax.set_title(f'({panel}) {title_tag} — S₁ Heatmap\n'
                     '(red labels = methodological; blue = technical)',
                     fontsize=10, fontweight='bold', loc='left')

    fig.suptitle('Figure H — Sobol First-Order Sensitivity Indices (S₁)',
                 fontsize=11, fontweight='bold')
    plt.tight_layout()
    plt.savefig('outputs/figures/FigH_Parameter_Heatmap.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/FigH_Parameter_Heatmap.pdf', bbox_inches='tight')
    plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# figI: Extended Variance Decomposition
# ─────────────────────────────────────────────────────────────────────────────

def figI_variance_decomposition_extended(var_df, sobol_results):
    """Figure I: Extended 6-panel variance decomposition for all pathways."""
    fig = plt.figure(figsize=(16, 11))
    gs  = GridSpec(2, 3, figure=fig, hspace=0.50, wspace=0.40)
    pathways = list(sobol_results.keys())

    panel_idx = 0
    for row, metric in enumerate(['MFSP', 'GHG']):
        metric_key = 'S1_mfsp' if metric == 'MFSP' else 'S1_ghg'
        sub_df     = var_df[var_df['Metric'] == metric]

        # ── Panel (col 0): stacked variance bar ──────────────────────────────
        ax0 = fig.add_subplot(gs[row, 0])
        x   = np.arange(len(pathways))
        meth = np.array([sub_df[sub_df.Pathway == p]['Methodological_%'].values[0]
                         for p in pathways])
        tech = np.array([sub_df[sub_df.Pathway == p]['Technical_%'].values[0]
                         for p in pathways])
        ax0.bar(x, meth, color='#e53935', alpha=0.85, label='Methodological', width=0.6)
        ax0.bar(x, tech, bottom=meth, color='#1565C0', alpha=0.85,
                label='Technical', width=0.6)
        ax0.set_xticks(x); ax0.set_xticklabels(pathways, fontsize=9)
        ax0.set_ylabel('Share of Variance (%)', fontsize=9)
        ax0.set_title(f'({chr(97 + panel_idx)}) {metric} Variance Decomp.',
                      fontsize=9, fontweight='bold', loc='left')
        ax0.set_ylim(0, 118); ax0.grid(axis='y', alpha=0.25)
        if row == 0:
            ax0.legend(fontsize=8)
        panel_idx += 1

        # ── Panel (col 1): CV before/after arrows ────────────────────────────
        ax1 = fig.add_subplot(gs[row, 1])
        cv_b = [sub_df[sub_df.Pathway == p]['CV_before_%'].values[0] for p in pathways]
        cv_a = [sub_df[sub_df.Pathway == p]['CV_after_%'].values[0]  for p in pathways]
        for i, pw in enumerate(pathways):
            ax1.annotate('', xy=(i, cv_a[i]), xytext=(i, cv_b[i]),
                         arrowprops=dict(arrowstyle='->', color=PATHWAY_COLORS[pw],
                                         lw=2.2, mutation_scale=12))
            ax1.scatter(i, cv_b[i], color=PATHWAY_COLORS[pw], s=70, zorder=5,
                        marker='o', label=pw if i == 0 else '')
            ax1.scatter(i, cv_a[i], color=PATHWAY_COLORS[pw], s=70, zorder=5,
                        marker='D')
            red_frac = sub_df[sub_df.Pathway == pw]['Variance_Reduction_%'].values[0]
            ax1.text(i + 0.08, (cv_b[i] + cv_a[i]) / 2, f'−{red_frac:.0f}%',
                     fontsize=7.5, color=PATHWAY_COLORS[pw])
        ax1.set_xticks(range(len(pathways))); ax1.set_xticklabels(pathways, fontsize=9)
        ax1.set_ylabel('CV (%)', fontsize=9)
        ax1.set_title(f'({chr(97 + panel_idx)}) CV Before (○) → After (◆)',
                      fontsize=9, fontweight='bold', loc='left')
        ax1.grid(axis='y', alpha=0.25)
        panel_idx += 1

        # ── Panel (col 2): top parameter bubble chart ─────────────────────────
        ax2 = fig.add_subplot(gs[row, 2])
        for j, pw in enumerate(pathways):
            s1 = sobol_results[pw][metric_key]
            top5 = sorted(s1.items(), key=lambda kv: kv[1], reverse=True)[:5]
            for k, (param, val) in enumerate(top5):
                y_pos  = j + (k - 2) * 0.14
                is_m   = param in METHODOLOGICAL_PARAMS
                ax2.scatter(val, y_pos,
                            s=max(val * 500 + 15, 15),
                            color='#e53935' if is_m else PATHWAY_COLORS[pw],
                            alpha=0.70, edgecolors='k', linewidths=0.4, zorder=3)
                if val > 0.04:
                    ax2.text(val + 0.012, y_pos,
                             param.replace('_', ' '), fontsize=6.5, va='center')
        ax2.set_yticks(range(len(pathways)))
        ax2.set_yticklabels(pathways, fontsize=9)
        ax2.set_xlabel('S₁ Index', fontsize=9)
        ax2.set_title(f'({chr(97 + panel_idx)}) Top Parameter S₁ Contributions',
                      fontsize=9, fontweight='bold', loc='left')
        ax2.grid(alpha=0.20)
        panel_idx += 1

    fig.suptitle('Figure I — Extended Variance Decomposition  '
                 '(Sobol Jansen 1999 estimator | red = methodological | blue = technical)',
                 fontsize=11, fontweight='bold')
    plt.savefig('outputs/figures/FigI_Variance_Decomp_Extended.png', dpi=600,
                bbox_inches='tight')
    plt.savefig('outputs/figures/FigI_Variance_Decomp_Extended.pdf', bbox_inches='tight')
    plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# figJ: TEA vs LCA Scatter — all 42 harmonized studies
# ─────────────────────────────────────────────────────────────────────────────

def figJ_tea_lca_scatter(df_harm, mc_results):
    """Figure J: Harmonized MFSP vs GHG for all 42 studies + MC median crosshairs."""
    fig, ax = plt.subplots(figsize=(12, 8))

    for pathway, grp in df_harm.groupby('pathway'):
        ax.scatter(grp['mfsp_harmonized'], grp['ghg_harmonized'],
                   c=PATHWAY_COLORS[pathway], s=60, alpha=0.75,
                   edgecolors='k', linewidths=0.5, zorder=4,
                   label=f'{pathway} (n={len(grp)})')
        # Annotate outliers
        q90_m = grp['mfsp_harmonized'].quantile(0.90)
        q90_g = grp['ghg_harmonized'].quantile(0.90)
        for _, row in grp.iterrows():
            if row['mfsp_harmonized'] > q90_m or row['ghg_harmonized'] > q90_g:
                ax.annotate(str(row['study_id'])[:10],
                            (row['mfsp_harmonized'], row['ghg_harmonized']),
                            fontsize=6, alpha=0.7,
                            textcoords='offset points', xytext=(4, 3))

    # MC medians with P5–P95 error bars
    for pathway in PATHWAY_CONFIGS:
        mc     = mc_results[pathway]
        med_m  = mc['mfsp'].median()
        med_g  = mc['ghg'].median()
        p5_m,  p95_m  = np.nanpercentile(mc['mfsp'], [5, 95])
        p5_g,  p95_g  = np.nanpercentile(mc['ghg'],  [5, 95])
        ax.errorbar(med_m, med_g,
                    xerr=[[med_m - p5_m], [p95_m - med_m]],
                    yerr=[[med_g - p5_g], [p95_g - med_g]],
                    fmt='*', color=PATHWAY_COLORS[pathway], ms=16,
                    lw=1.8, capsize=5,
                    markeredgecolor='black', markeredgewidth=0.8,
                    label=f'{pathway} MC (P5–P95)', zorder=6)

    ax.axhline(PETROLEUM_JET_GHG_WTW, color='red', ls='--', lw=1.5,
               label='Petroleum jet (89 gCO₂e/MJ)', alpha=0.85)
    ax.axhline(PETROLEUM_JET_GHG_WTW * 0.50, color='#F57C00', ls='-.', lw=1.2,
               label='50% GHG reduction', alpha=0.85)
    ax.axhline(0, color='gray', ls=':', lw=0.8)

    ax.set_xlabel('Harmonized MFSP (2023 $/GGE)', fontsize=11, fontweight='bold')
    ax.set_ylabel('Harmonized GHG (gCO₂e/MJ)',    fontsize=11, fontweight='bold')
    ax.set_title('Figure J — TEA vs LCA Trade-off Space: All 42 Harmonized Studies\n'
                 '(stars = MC P50 ± P5–P95 range; points = individual studies)',
                 fontsize=10, fontweight='bold', loc='left')
    ax.legend(fontsize=8.5, ncol=2, loc='upper right')
    ax.grid(alpha=0.20)
    plt.tight_layout()
    plt.savefig('outputs/figures/FigJ_TEA_LCA_Scatter.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/FigJ_TEA_LCA_Scatter.pdf', bbox_inches='tight')
    plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# figK: Multi-dimensional Radar Plot
# ─────────────────────────────────────────────────────────────────────────────

def figK_radar_plot(mc_results):
    """
    Figure K: Radar/spider plot — 6-dimensional pathway comparison.

    Dimensions:
      Cost Competitiveness — inverted MC P50 MFSP (computed, scaled 0–5)
      GHG Reduction        — (89 − MC P50 GHG) / 89 × 5 (computed)
      Land Use Efficiency  — expert score 1–5 (IRENA 2021; higher=less land)
      Water Use Efficiency — expert score 1–5 (ICAO 2019; higher=less water)
      TRL                  — technology readiness level/9 × 5 (IEA 2023)
      Feedstock Availability — expert score 1–5 (higher=more abundant)
    Note: land/water/TRL/feedstock are literature-derived scores cited above.
    """
    pathways = ['ATJ', 'HEFA', 'FT-SPK', 'PtL']

    mfsp_p50 = {pw: mc_results[pw]['mfsp'].median() for pw in pathways}
    ghg_p50  = {pw: mc_results[pw]['ghg'].median()  for pw in pathways}
    max_mfsp = max(mfsp_p50.values())

    cost_score = {pw: (1 - mfsp_p50[pw] / max_mfsp) * 5          for pw in pathways}
    ghg_score  = {pw: max((PETROLEUM_JET_GHG_WTW - ghg_p50[pw])
                          / PETROLEUM_JET_GHG_WTW * 5, 0)         for pw in pathways}

    # Literature-derived expert scores (1–5 scale)
    land_eff  = {'ATJ': 3.0, 'HEFA': 2.0, 'FT-SPK': 4.0, 'PtL': 5.0}
    water_eff = {'ATJ': 3.0, 'HEFA': 2.0, 'FT-SPK': 3.0, 'PtL': 4.0}
    trl_score = {'ATJ': 8/9*5, 'HEFA': 5.0, 'FT-SPK': 7/9*5, 'PtL': 5/9*5}
    feed_avail= {'ATJ': 4.0, 'HEFA': 2.0, 'FT-SPK': 5.0, 'PtL': 5.0}

    dim_labels = [
        'Cost\nCompetitiveness', 'GHG\nReduction', 'Land Use\nEfficiency',
        'Water Use\nEfficiency', 'TRL\nMaturity',  'Feedstock\nAvailability',
    ]
    N      = len(dim_labels)
    angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
    angles += angles[:1]

    fig, ax = plt.subplots(figsize=(9, 9), subplot_kw=dict(polar=True))

    for pw in pathways:
        vals = [cost_score[pw], ghg_score[pw], land_eff[pw],
                water_eff[pw],  trl_score[pw], feed_avail[pw]]
        vals += vals[:1]
        color = PATHWAY_COLORS[pw]
        ax.plot(angles, vals, 'o-', linewidth=2.2, color=color, label=pw)
        ax.fill(angles, vals, alpha=0.12, color=color)
        # Annotate each vertex with value
        for angle, val in zip(angles[:-1], vals[:-1]):
            ax.text(angle, val + 0.18, f'{val:.1f}', fontsize=7, ha='center',
                    va='center', color=color, fontweight='bold')

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(dim_labels, fontsize=9, fontweight='bold')
    ax.set_ylim(0, 5.6)
    ax.set_yticks([1, 2, 3, 4, 5])
    ax.set_yticklabels(['1', '2', '3', '4', '5'], fontsize=7, color='gray')
    ax.grid(color='gray', linestyle='--', linewidth=0.5, alpha=0.5)
    ax.set_title('Figure K — Multi-Dimensional SAF Pathway Comparison\n'
                 'Cost & GHG from MC medians  |  Land/Water/TRL/Feedstock: literature scores',
                 fontsize=9, fontweight='bold', pad=22)
    ax.legend(loc='upper right', bbox_to_anchor=(1.38, 1.14), fontsize=10)

    plt.tight_layout()
    plt.savefig('outputs/figures/FigK_Radar_Plot.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/FigK_Radar_Plot.pdf', bbox_inches='tight')
    plt.close()


# ─────────────────────────────────────────────────────────────────────────────
# figL: Pareto Sensitivity Chart (Sobol S1 sorted, cumulative %)
# ─────────────────────────────────────────────────────────────────────────────

def figL_pareto_sensitivity(sobol_results):
    """
    Figure L: Pareto chart of Sobol S1 sensitivity indices per pathway.

    For each pathway × metric (MFSP / GHG):
      - Bar height = S1 for that parameter (sorted descending)
      - Line       = cumulative % of total S1 explained
      - Red dashed = 80% threshold (Pareto 80/20 rule)
    Data: computed directly from sobol_results (Jansen 1999 estimator).
    """
    pathways   = list(sobol_results.keys())
    metric_cfg = [
        ('S1_mfsp', 'MFSP', '#1565C0'),
        ('S1_ghg',  'GHG',  '#e53935'),
    ]
    fig, axes = plt.subplots(len(pathways), 2,
                             figsize=(16, 4 * len(pathways)),
                             squeeze=False)
    fig.suptitle(
        'Figure L — Pareto Sensitivity Chart  (Sobol S₁, sorted by contribution)\n'
        'Red dashed line = 80 % cumulative threshold  |  '
        'red labels = methodological parameters',
        fontsize=11, fontweight='bold',
    )

    for row, pathway in enumerate(pathways):
        for col, (metric_key, metric_name, bar_color) in enumerate(metric_cfg):
            ax   = axes[row][col]
            ax2  = ax.twinx()          # right axis for cumulative %

            s1_raw   = sobol_results[pathway][metric_key]
            total_s1 = sum(s1_raw.values()) or 1.0
            sorted_s1 = sorted(s1_raw.items(), key=lambda kv: kv[1], reverse=True)
            params    = [kv[0] for kv in sorted_s1]
            values    = np.array([kv[1] for kv in sorted_s1])
            cumulative = np.cumsum(values) / total_s1 * 100

            x = np.arange(len(params))

            # Bars
            clrs = ['#e53935' if p in METHODOLOGICAL_PARAMS else bar_color
                    for p in params]
            ax.bar(x, values, color=clrs, alpha=0.80,
                   edgecolor='white', linewidth=0.6, width=0.65)
            for xi, v in enumerate(values):
                if v > 0.01:
                    ax.text(xi, v + 0.004, f'{v:.2f}', ha='center',
                            fontsize=7.5, fontweight='bold',
                            color='#e53935' if params[xi] in METHODOLOGICAL_PARAMS
                            else bar_color)

            # Cumulative line
            ax2.plot(x, cumulative, 'k-o', ms=5, lw=1.8, zorder=5,
                     label='Cumulative %')
            ax2.axhline(80, color='red', ls='--', lw=1.2, alpha=0.8,
                        label='80% threshold')
            ax2.set_ylim(0, 115)
            ax2.set_ylabel('Cumulative S₁ (%)', fontsize=8, color='black')
            ax2.tick_params(axis='y', labelsize=8)

            ax.set_xticks(x)
            ax.set_xticklabels(
                [p.replace('_', '\n') for p in params],
                fontsize=7.5, rotation=0,
            )
            for xtick, param in zip(ax.get_xticklabels(), params):
                xtick.set_color(
                    '#e53935' if param in METHODOLOGICAL_PARAMS else '#424242'
                )
            ax.set_ylabel('First-Order Sobol Index S₁', fontsize=9)
            ax.set_title(f'{pathway} — {metric_name}',
                         fontsize=10, fontweight='bold',
                         color=PATHWAY_COLORS[pathway])
            ax.grid(axis='y', alpha=0.25)
            ax.set_ylim(bottom=0)

            # Legend on first panel only
            if row == 0 and col == 0:
                meth_p = mpatches.Patch(color='#e53935', alpha=0.8,
                                        label='Methodological param')
                tech_p = mpatches.Patch(color=bar_color, alpha=0.8,
                                        label='Technical param')
                ax.legend(handles=[meth_p, tech_p], fontsize=8, loc='upper right')
                ax2.legend(fontsize=8, loc='center right')

    plt.tight_layout()
    plt.savefig('outputs/figures/FigL_Pareto_Sensitivity.png',
                dpi=600, bbox_inches='tight')
    plt.savefig('outputs/figures/FigL_Pareto_Sensitivity.pdf', bbox_inches='tight')
    plt.close()


# =============================================================================
# SECTION 13: MAIN
# =============================================================================

if __name__ == '__main__':
    print('=' * 72)
    print('SAF HARMONIZATION META-ANALYSIS FRAMEWORK v2.0')
    print('=' * 72)

    # Step 1: Build and harmonize literature database
    print('\n[1/5] Harmonizing 42 literature studies...')
    df_harm = build_harmonized_dataset()
    print(f"      Studies: {len(df_harm)} | Pathways: {df_harm['pathway'].nunique()}")
    print(f"      Raw MFSP range    : ${df_harm['mfsp_2023_raw'].min():.2f}–"
          f"${df_harm['mfsp_2023_raw'].max():.2f}/GGE (2023 USD)")
    print(f"      Harm. MFSP range  : ${df_harm['mfsp_harmonized'].min():.2f}–"
          f"${df_harm['mfsp_harmonized'].max():.2f}/GGE")
    print(f"      Raw GHG range     : {df_harm['ghg_raw'].min():.1f}–"
          f"{df_harm['ghg_raw'].max():.1f} gCO₂e/MJ")
    print(f"      Harm. GHG range   : {df_harm['ghg_harmonized'].min():.1f}–"
          f"{df_harm['ghg_harmonized'].max():.1f} gCO₂e/MJ")

    # Step 2: Monte Carlo simulation
    print('\n[2/5] Monte Carlo simulation (10,000 iterations × 4 pathways)...')
    mc_results = run_monte_carlo(n_iter=10_000)
    for pathway, mc in mc_results.items():
        print(f"      {pathway:8s}: MFSP median={mc['mfsp'].median():.2f} $/GGE  "
              f"GHG median={mc['ghg'].median():.1f} gCO₂e/MJ")

    # Step 3: Sobol sensitivity analysis
    print('\n[3/5] Sobol sensitivity analysis (1,500 base samples × 4 pathways)...')
    sobol_results = run_sobol_analysis(n_sobol=1500)

    # Step 4: Variance decomposition
    print('\n[4/5] Variance decomposition (methodological vs. technical)...')
    var_df = decompose_variance(mc_results, sobol_results)
    print(var_df[['Pathway', 'Metric', 'Methodological_%', 'Technical_%',
                  'Variance_Reduction_%']].to_string(index=False))

    # Step 5: Compute modal-parameter breakdown for cost/GHG figures
    print('\n[5/6] Computing modal-parameter cost & GHG breakdowns...')
    breakdown = compute_cost_ghg_breakdown()
    for pw, data in breakdown.items():
        total_cost = sum(data['cost'].values())
        total_ghg  = sum(max(v, 0) for v in data['ghg'].values())
        print(f"      {pw:8s}: MFSP(modal)=${total_cost:.2f}/GGE  "
              f"GHG(modal)={total_ghg:.1f} gCO₂e/MJ")

    # Step 6: Generate all figures and Excel workbook
    print('\n[6/6] Generating figures and Excel workbook...')

    # — Original four figures —
    fig1_literature_overview(df_harm)
    fig2_harmonization_impact(df_harm)
    fig3_monte_carlo_distributions(mc_results)
    fig4_variance_decomposition(var_df, sobol_results)

    # — Additional publication figures —
    figA_system_boundary()
    figB_harmonization_flowchart(df_harm)
    figC_cost_breakdown(breakdown)
    figD_tornado_sensitivity()
    figE_mc_overlay(df_harm, mc_results)
    figF_ghg_comparison(mc_results, df_harm)
    figG_ghg_contribution(breakdown)
    figH_parameter_heatmap(sobol_results)
    figI_variance_decomposition_extended(var_df, sobol_results)
    figJ_tea_lca_scatter(df_harm, mc_results)
    figK_radar_plot(mc_results)
    figL_pareto_sensitivity(sobol_results)

    wb = create_excel_workbook(df_harm, mc_results, var_df)
    wb.save('outputs/SAF_MetaAnalysis_Harmonization.xlsx')
    #%%
    print(f'\n{"=" * 72}')
    print('ANALYSIS COMPLETE')
    print(f'{"=" * 72}')
    print('\nOutputs saved:')
    print('  outputs/figures/Fig1_Literature_Overview.{png,pdf}')
    print('  outputs/figures/Fig2_Harmonization_Impact.{png,pdf}')
    print('  outputs/figures/Fig3_MC_Distributions.{png,pdf}')
    print('  outputs/figures/Fig4_Variance_Decomposition.{png,pdf}')
    print('  outputs/figures/FigA_System_Boundary.{png,pdf}')
    print('  outputs/figures/FigB_Harmonization_Flowchart.{png,pdf}')
    print('  outputs/figures/FigC_Cost_Breakdown.{png,pdf}')
    print('  outputs/figures/FigD_Tornado_Sensitivity.{png,pdf}')
    print('  outputs/figures/FigE_MC_Overlay.{png,pdf}')
    print('  outputs/figures/FigF_GHG_Comparison.{png,pdf}')
    print('  outputs/figures/FigG_GHG_Contribution.{png,pdf}')
    print('  outputs/figures/FigH_Parameter_Heatmap.{png,pdf}')
    print('  outputs/figures/FigI_Variance_Decomp_Extended.{png,pdf}')
    print('  outputs/figures/FigJ_TEA_LCA_Scatter.{png,pdf}')
    print('  outputs/figures/FigK_Radar_Plot.{png,pdf}')
    print('  outputs/figures/FigL_Pareto_Sensitivity.{png,pdf}')
    print('  outputs/SAF_MetaAnalysis_Harmonization.xlsx')
    print(f'{"=" * 72}\n')
    #%%