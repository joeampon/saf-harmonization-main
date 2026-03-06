"""
SAF Harmonization Framework - Computational Analysis
Quantifying methodological variability in sustainable aviation fuel assessments
Data sources: Tao et al. 2017, Han et al. 2017, GREET Model 2023
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
import os
import warnings
warnings.filterwarnings('ignore')

plt.rcParams['font.family'] = 'DejaVu Sans'
plt.rcParams['font.size'] = 10
plt.rcParams['axes.linewidth'] = 1.2
sns.set_palette("husl")

os.makedirs('outputs', exist_ok=True)

class BaselineData:
    """Technoeconomic baseline parameters for SAF production."""
    FEEDSTOCK_INPUT_TPD = 2000
    CAPACITY_FACTOR = 0.90
    PLANT_LIFETIME = 30
    DISCOUNT_RATE = 0.10
    FEEDSTOCK_COST = 80.00
    FEEDSTOCK_TRANSPORT_COST_PER_TONNE = 12.00
    CAPITAL_COST_2011 = 420_000_000
    ETHANOL_YIELD_GAL_PER_TONNE = 79
    JET_YIELD_FROM_ETHANOL = 0.71
    DIESEL_COPROD_FROM_ETHANOL = 0.15
    GASOLINE_COPROD_FROM_ETHANOL = 0.14
    JET_LHV_BTU_PER_GAL = 120200
    GASOLINE_LHV_BTU_PER_GAL = 112194
    DIESEL_LHV_BTU_PER_GAL = 129488
    FIXED_OPEX_PERCENT = 0.04
    VARIABLE_OPEX_PERCENT = 0.03

class LCAData:
    """Life-cycle assessment baseline parameters for GHG emissions."""
    BASELINE_GHG_GCO2E_PER_MJ = 24
    PETROLEUM_JET_BASELINE = 89
    FEEDSTOCK_PRODUCTION_GCO2E_MJ = 15
    NATURAL_GAS_USE_MJ_PER_MJ_FUEL = 0.15
    ELECTRICITY_USE_KWH_PER_MJ_FUEL = 0.02
    US_GRID_ELECTRICITY_GCO2E_KWH = 450
    NATURAL_GAS_GCO2E_MJ = 11

def calculate_technoeconomic_metrics(
    feedstock_cost=BaselineData.FEEDSTOCK_COST,
    feedstock_transport_cost_per_tonne=BaselineData.FEEDSTOCK_TRANSPORT_COST_PER_TONNE,
    discount_rate=BaselineData.DISCOUNT_RATE,
    capacity_factor=BaselineData.CAPACITY_FACTOR,
    capital_cost=BaselineData.CAPITAL_COST_2011,
    include_feedstock_transport=True,
    include_distribution=False
):
    """Calculate minimum fuel selling price (MFSP) under specified conditions.
    
    Args:
        feedstock_cost: Feedstock purchase price ($/tonne).
        feedstock_transport_cost_per_tonne: Transport cost per tonne.
        discount_rate: Plant capital cost discount rate (%).
        capacity_factor: Plant utilization rate (0-1).
        capital_cost: Initial plant capital investment ($).
        include_feedstock_transport: Include transport surcharge if True.
        include_distribution: Include distribution cost if True.
    
    Returns:
        dict: MFSP value in $/GGE.
    """
    annual_feedstock_tonnes = BaselineData.FEEDSTOCK_INPUT_TPD * 365 * capacity_factor
    # Convert feedstock to intermediate ethanol and co-products (diesel, gasoline)
    annual_ethanol_gal = annual_feedstock_tonnes * BaselineData.ETHANOL_YIELD_GAL_PER_TONNE
    annual_jet_gal = annual_ethanol_gal * BaselineData.JET_YIELD_FROM_ETHANOL
    annual_diesel_gal = annual_ethanol_gal * BaselineData.DIESEL_COPROD_FROM_ETHANOL
    annual_gasoline_gal = annual_ethanol_gal * BaselineData.GASOLINE_COPROD_FROM_ETHANOL
    
    # Normalize fuel outputs to gasoline equivalent (GGE) for cost allocation
    jet_gge = annual_jet_gal * (BaselineData.JET_LHV_BTU_PER_GAL / BaselineData.GASOLINE_LHV_BTU_PER_GAL)
    diesel_gge = annual_diesel_gal * (BaselineData.DIESEL_LHV_BTU_PER_GAL / BaselineData.GASOLINE_LHV_BTU_PER_GAL)
    total_fuel_gge = jet_gge + diesel_gge + annual_gasoline_gal
    
    # Capital recovery factor for 30-year plant life
    crf = (discount_rate * (1 + discount_rate)**BaselineData.PLANT_LIFETIME) / \
          ((1 + discount_rate)**BaselineData.PLANT_LIFETIME - 1)
    
    # Annualize capital and operating costs
    annual_capital_charge = capital_cost * crf
    fixed_opex = capital_cost * BaselineData.FIXED_OPEX_PERCENT
    variable_opex = capital_cost * BaselineData.VARIABLE_OPEX_PERCENT
    # Feedstock always charged; transport is optional
    feedstock_purchase_cost = annual_feedstock_tonnes * feedstock_cost
    transport_surcharge = annual_feedstock_tonnes * feedstock_transport_cost_per_tonne if include_feedstock_transport else 0
    feedstock_annual_cost = feedstock_purchase_cost + transport_surcharge
    distribution_cost = total_fuel_gge * 0.15 if include_distribution else 0
    
    total_annual_cost = annual_capital_charge + fixed_opex + variable_opex + feedstock_annual_cost + distribution_cost
    mfsp_per_gge = total_annual_cost / total_fuel_gge
    
    return {'MFSP_$/GGE': mfsp_per_gge}

def calculate_lca_metrics(
    allocation_method='energy',
    include_iluc=False,
    electricity_grid_intensity=LCAData.US_GRID_ELECTRICITY_GCO2E_KWH
):
    """Calculate life-cycle GHG emissions per MJ of fuel.
    
    Args:
        allocation_method: Co-product allocation method (energy, mass, economic, system_expansion).
        include_iluc: Add indirect land-use change emissions if True.
        electricity_grid_intensity: Grid electricity carbon intensity (gCO2e/kWh).
    
    Returns:
        dict: GHG emissions and reduction percentage relative to petroleum baseline.
    """
    # Sum lifecycle emissions: feedstock production + conversion (natural gas + electricity)
    feedstock_ghg = LCAData.FEEDSTOCK_PRODUCTION_GCO2E_MJ
    ng_emissions = LCAData.NATURAL_GAS_USE_MJ_PER_MJ_FUEL * LCAData.NATURAL_GAS_GCO2E_MJ
    elec_emissions = LCAData.ELECTRICITY_USE_KWH_PER_MJ_FUEL * electricity_grid_intensity
    conversion_ghg = ng_emissions + elec_emissions
    total_ghg_before_allocation = feedstock_ghg + conversion_ghg
    
    # Apply co-product allocation method (varies from 0.60 to 0.75 of total emissions assigned to jet fuel)
    allocation_factors = {
        'energy': 0.71,
        'mass': 0.75,
        'economic': 0.69,
        'system_expansion': 0.60
    }
    allocation_factor = allocation_factors.get(allocation_method, 0.71)
    
    allocated_ghg = total_ghg_before_allocation * allocation_factor
    # Optional: Add indirect land-use change penalty
    if include_iluc:
        allocated_ghg += 5
    
    ghg_reduction_percent = ((LCAData.PETROLEUM_JET_BASELINE - allocated_ghg) / 
                            LCAData.PETROLEUM_JET_BASELINE * 100)
    
    return {
        'GHG_gCO2e/MJ': allocated_ghg,
        'GHG_Reduction_%': ghg_reduction_percent
    }

def run_full_analysis():
    """Execute 14-scenario sensitivity analysis across methodological variations.
    
    Returns:
        DataFrame: Scenario results with technoeconomic and LCA metrics.
    """
    scenarios = []
    
    # Baseline scenario with default parameters from both data classes
    baseline = {
        'Scenario': 'Baseline',
        'Short_Name': 'BASE',
        'System_Boundary': 'Well-to-Gate',
        'Allocation': 'Energy',
        'Discount_Rate_%': 10,
        'Feedstock_Cost_$/tonne': 80,
        'Capacity_Factor_%': 90,
        'Include_ILUC': False,
        'Grid_Intensity_gCO2e/kWh': 450,
        'Category': 'Baseline'
    }
    tec = calculate_technoeconomic_metrics()
    lca = calculate_lca_metrics()
    baseline.update({
        'MFSP_$/GGE': tec['MFSP_$/GGE'],
        'GHG_gCO2e/MJ': lca['GHG_gCO2e/MJ'],
        'GHG_Reduction_%': lca['GHG_Reduction_%'],
        'MFSP_Change_%': 0,
        'GHG_Change_%': 0
    })
    scenarios.append(baseline)
    
    # Test 12 sensitivity scenarios across 4 categories: boundary, allocation, economics, environmental
    scenario_configs = [
        ('S1: Gate-to-Gate', 'S1', 'Gate-to-Gate', 'Boundary', 
         {'include_feedstock_transport': False}, {}),
        ('S2: Well-to-Pump', 'S2', 'Well-to-Pump', 'Boundary',
         {'include_distribution': True}, {}),
        ('S3: Mass Allocation', 'S3', 'Well-to-Gate', 'Allocation',
         {}, {'allocation_method': 'mass'}),
        ('S4: Economic Allocation', 'S4', 'Well-to-Gate', 'Allocation',
         {}, {'allocation_method': 'economic'}),
        ('S5: System Expansion', 'S5', 'Well-to-Gate', 'Allocation',
         {}, {'allocation_method': 'system_expansion'}),
        ('S6: Discount 5%', 'S6', 'Well-to-Gate', 'Economics',
         {'discount_rate': 0.05}, {}),
        ('S7: Discount 15%', 'S7', 'Well-to-Gate', 'Economics',
         {'discount_rate': 0.15}, {}),
        ('S8: Feedstock $40', 'S8', 'Well-to-Gate', 'Economics',
         {'feedstock_cost': 40}, {}),
        ('S9: Feedstock $120', 'S9', 'Well-to-Gate', 'Economics',
         {'feedstock_cost': 120}, {}),
        ('S10: With ILUC', 'S10', 'Well-to-Gate', 'Environmental',
         {}, {'include_iluc': True}),
        ('S11: Renewable Elec', 'S11', 'Well-to-Gate', 'Environmental',
         {}, {'electricity_grid_intensity': 50}),
        ('S12: Capacity 80%', 'S12', 'Well-to-Gate', 'Operational',
         {'capacity_factor': 0.80}, {}),
    ]
    
    for name, short, boundary, category, tec_kwargs, lca_kwargs in scenario_configs:
        s = baseline.copy()
        s.update({
            'Scenario': name,
            'Short_Name': short,
            'System_Boundary': boundary,
            'Category': category
        })

        # Update metadata fields to match scenario parameters for transparency
        if 'discount_rate' in tec_kwargs:
            s['Discount_Rate_%'] = tec_kwargs['discount_rate'] * 100
        if 'feedstock_cost' in tec_kwargs:
            s['Feedstock_Cost_$/tonne'] = tec_kwargs['feedstock_cost']
        if 'capacity_factor' in tec_kwargs:
            s['Capacity_Factor_%'] = tec_kwargs['capacity_factor'] * 100

        # Map allocation method keywords to display labels
        allocation_labels = {
            'energy': 'Energy',
            'mass': 'Mass',
            'economic': 'Economic',
            'system_expansion': 'System Expansion'
        }
        if 'allocation_method' in lca_kwargs:
            method = lca_kwargs['allocation_method']
            s['Allocation'] = allocation_labels.get(method, method.replace('_', ' ').title())
        if 'include_iluc' in lca_kwargs:
            s['Include_ILUC'] = lca_kwargs['include_iluc']
        if 'electricity_grid_intensity' in lca_kwargs:
            s['Grid_Intensity_gCO2e/kWh'] = lca_kwargs['electricity_grid_intensity']
        
        if tec_kwargs:
            tec = calculate_technoeconomic_metrics(**tec_kwargs)
            s['MFSP_$/GGE'] = tec['MFSP_$/GGE']
            s['MFSP_Change_%'] = ((s['MFSP_$/GGE'] - baseline['MFSP_$/GGE']) / 
                                 baseline['MFSP_$/GGE'] * 100)
        
        # Calculate deviations from baseline for LCA metrics
        if lca_kwargs:
            lca = calculate_lca_metrics(**lca_kwargs)
            s['GHG_gCO2e/MJ'] = lca['GHG_gCO2e/MJ']
            s['GHG_Reduction_%'] = lca['GHG_Reduction_%']
            s['GHG_Change_%'] = ((s['GHG_gCO2e/MJ'] - baseline['GHG_gCO2e/MJ']) / 
                                baseline['GHG_gCO2e/MJ'] * 100)
        
        scenarios.append(s)
    
    # Pessimistic scenario: High discount rate + high feedstock cost + economic allocation + ILUC
    s13 = baseline.copy()
    s13.update({
        'Scenario': 'S13: Pessimistic',
        'Short_Name': 'S13',
        'Discount_Rate_%': 15,
        'Feedstock_Cost_$/tonne': 120,
        'Allocation': 'Economic',
        'Include_ILUC': True,
        'Category': 'Combined'
    })
    tec = calculate_technoeconomic_metrics(discount_rate=0.15, feedstock_cost=120)
    lca = calculate_lca_metrics(allocation_method='economic', include_iluc=True)
    s13['MFSP_$/GGE'] = tec['MFSP_$/GGE']
    s13['GHG_gCO2e/MJ'] = lca['GHG_gCO2e/MJ']
    s13['MFSP_Change_%'] = ((s13['MFSP_$/GGE'] - baseline['MFSP_$/GGE']) / baseline['MFSP_$/GGE'] * 100)
    s13['GHG_Change_%'] = ((s13['GHG_gCO2e/MJ'] - baseline['GHG_gCO2e/MJ']) / baseline['GHG_gCO2e/MJ'] * 100)
    scenarios.append(s13)
    
    # Optimistic scenario: Low discount rate + low feedstock cost + system expansion allocation + renewable electricity
    s14 = baseline.copy()
    s14.update({
        'Scenario': 'S14: Optimistic',
        'Short_Name': 'S14',
        'Discount_Rate_%': 5,
        'Feedstock_Cost_$/tonne': 40,
        'Allocation': 'System Expansion',
        'Grid_Intensity_gCO2e/kWh': 50,
        'Category': 'Combined'
    })
    tec = calculate_technoeconomic_metrics(discount_rate=0.05, feedstock_cost=40)
    lca = calculate_lca_metrics(allocation_method='system_expansion', electricity_grid_intensity=50)
    s14['MFSP_$/GGE'] = tec['MFSP_$/GGE']
    s14['GHG_gCO2e/MJ'] = lca['GHG_gCO2e/MJ']
    s14['MFSP_Change_%'] = ((s14['MFSP_$/GGE'] - baseline['MFSP_$/GGE']) / baseline['MFSP_$/GGE'] * 100)
    s14['GHG_Change_%'] = ((s14['GHG_gCO2e/MJ'] - baseline['GHG_gCO2e/MJ']) / baseline['GHG_gCO2e/MJ'] * 100)
    scenarios.append(s14)
    
    return pd.DataFrame(scenarios)

def create_main_results_plot(df):
    """Generate cost and emissions bar charts for all scenarios.
    
    Args:
        df: Scenario results DataFrame.
    
    Outputs:
        PNG and PDF figures saved to outputs/ directory.
    """
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
    
    colors_cost = ['#1f77b4' if 'BASE' in s else '#ff7f0e' if 'S1' in s or 'S2' in s 
                   else '#2ca02c' if any(x in s for x in ['S6', 'S7', 'S8', 'S9'])
                   else '#d62728' if 'S13' in s or 'S14' in s else '#9467bd' 
                   for s in df['Short_Name']]
    
    ax1.bar(range(len(df)), df['MFSP_$/GGE'], color=colors_cost, edgecolor='black', linewidth=0.8)
    ax1.axhline(y=df.iloc[0]['MFSP_$/GGE'], color='gray', linestyle='--', linewidth=1, alpha=0.7)
    ax1.set_ylabel('MFSP ($/GGE)', fontsize=11, fontweight='bold')
    ax1.set_xlabel('Scenario', fontsize=11, fontweight='bold')
    ax1.set_xticks(range(len(df)))
    ax1.set_xticklabels(df['Short_Name'], rotation=45, ha='right')
    ax1.set_title('(a) Minimum Fuel Selling Price', fontsize=12, fontweight='bold', loc='left')
    ax1.grid(axis='y', alpha=0.3)
    ax1.set_ylim(0, df['MFSP_$/GGE'].max() * 1.15)
    
    for i, v in enumerate(df['MFSP_$/GGE']):
        ax1.text(i, v + 0.05, f'${v:.2f}', ha='center', va='bottom', fontsize=8)
    
    colors_ghg = ['#1f77b4' if 'BASE' in s else '#ff7f0e' if any(x in s for x in ['S3', 'S4', 'S5'])
                  else '#2ca02c' if 'S10' in s or 'S11' in s 
                  else '#d62728' if 'S13' in s or 'S14' in s else '#9467bd'
                  for s in df['Short_Name']]
    
    ax2.bar(range(len(df)), df['GHG_gCO2e/MJ'], color=colors_ghg, edgecolor='black', linewidth=0.8)
    ax2.axhline(y=df.iloc[0]['GHG_gCO2e/MJ'], color='gray', linestyle='--', linewidth=1, alpha=0.7)
    ax2.axhline(y=89, color='red', linestyle=':', linewidth=1.5, alpha=0.7)
    ax2.set_ylabel('GHG Emissions (gCO₂e/MJ)', fontsize=11, fontweight='bold')
    ax2.set_xlabel('Scenario', fontsize=11, fontweight='bold')
    ax2.set_xticks(range(len(df)))
    ax2.set_xticklabels(df['Short_Name'], rotation=45, ha='right')
    ax2.set_title('(b) Greenhouse Gas Emissions', fontsize=12, fontweight='bold', loc='left')
    ax2.grid(axis='y', alpha=0.3)
    ax2.set_ylim(0, 95)
    
    for i, v in enumerate(df['GHG_gCO2e/MJ']):
        ax2.text(i, v + 1, f'{v:.1f}', ha='center', va='bottom', fontsize=8)
    
    plt.tight_layout()
    plt.savefig('outputs/Figure_1_Main_Results.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/Figure_1_Main_Results.pdf', bbox_inches='tight')
    plt.close()

def create_harmonization_plot(df):
    """Generate violin plots showing variance reduction from methodological harmonization.
    
    Args:
        df: Scenario results DataFrame.
    
    Outputs:
        PNG and PDF figures saved to outputs/ directory.
    """
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
    
    before_cost = df['MFSP_$/GGE']
    harmonized_mask = df['Category'].isin(['Baseline', 'Economics', 'Environmental', 'Operational', 'Combined'])
    after_cost = df[harmonized_mask]['MFSP_$/GGE']
    
    parts1 = ax1.violinplot([before_cost, after_cost], positions=[1, 2], 
                            showmeans=True, showmedians=True, widths=0.7)
    for pc in parts1['bodies']:
        pc.set_facecolor('#2ca02c')
        pc.set_alpha(0.7)
    
    ax1.scatter([1]*len(before_cost), before_cost, alpha=0.5, s=50, color='darkblue', zorder=3)
    ax1.scatter([2]*len(after_cost), after_cost, alpha=0.5, s=50, color='darkred', zorder=3)
    ax1.set_xticks([1, 2])
    ax1.set_xticklabels(['Before\nHarmonization', 'After\nHarmonization'])
    ax1.set_ylabel('MFSP ($/GGE)', fontsize=11, fontweight='bold')
    ax1.set_title('(a) Cost Variance Reduction', fontsize=12, fontweight='bold', loc='left')
    ax1.grid(axis='y', alpha=0.3)
    
    cv_before_cost = (before_cost.std() / before_cost.mean()) * 100
    cv_after_cost = (after_cost.std() / after_cost.mean()) * 100
    reduction_cost = ((cv_before_cost - cv_after_cost) / cv_before_cost) * 100
    
    before_ghg = df['GHG_gCO2e/MJ']
    after_ghg = df[harmonized_mask]['GHG_gCO2e/MJ']
    
    parts2 = ax2.violinplot([before_ghg, after_ghg], positions=[1, 2],
                            showmeans=True, showmedians=True, widths=0.7)
    for pc in parts2['bodies']:
        pc.set_facecolor('#ff7f0e')
        pc.set_alpha(0.7)
    
    ax2.scatter([1]*len(before_ghg), before_ghg, alpha=0.5, s=50, color='darkgreen', zorder=3)
    ax2.scatter([2]*len(after_ghg), after_ghg, alpha=0.5, s=50, color='darkred', zorder=3)
    ax2.set_xticks([1, 2])
    ax2.set_xticklabels(['Before\nHarmonization', 'After\nHarmonization'])
    ax2.set_ylabel('GHG (gCO₂e/MJ)', fontsize=11, fontweight='bold')
    ax2.set_title('(b) GHG Variance Reduction', fontsize=12, fontweight='bold', loc='left')
    ax2.grid(axis='y', alpha=0.3)
    
    cv_before_ghg = (before_ghg.std() / before_ghg.mean()) * 100
    cv_after_ghg = (after_ghg.std() / after_ghg.mean()) * 100
    reduction_ghg = ((cv_before_ghg - cv_after_ghg) / cv_before_ghg) * 100
    
    plt.tight_layout()
    plt.savefig('outputs/Figure_2_Harmonization_Impact.png', dpi=600, bbox_inches='tight')
    plt.savefig('outputs/Figure_2_Harmonization_Impact.pdf', bbox_inches='tight')
    plt.close()

def create_excel_workbook(df):
    """Create multi-sheet Excel workbook with results and data sources.
    
    Args:
        df: Scenario results DataFrame.
    
    Returns:
        Workbook: openpyxl Workbook object with Data Sources, Results, and Summary sheets.
    """
    wb = Workbook()
    wb.remove(wb.active)
    
    ws_sources = wb.create_sheet("Data Sources")
    ws_sources['A1'] = "VERIFIED DATA SOURCES"
    ws_sources['A1'].font = Font(bold=True, size=14)
    
    sources = [
        ["Source", "Citation", "Data Used", "DOI"],
        ["Tao et al. 2017", "Green Chemistry", "TEA Parameters", "10.1039/C6GC02800D"],
        ["Han et al. 2017", "Biotechnol Biofuels", "LCA Emissions", "10.1186/s13068-017-0698-z"],
        ["GREET 2023", "Argonne National Lab", "Emission Factors", "greet.es.anl.gov"],
    ]
    
    for row_idx, row_data in enumerate(sources, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_sources.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
    
    for col in ['A', 'B', 'C', 'D']:
        ws_sources.column_dimensions[col].width = 25
    
    ws_results = wb.create_sheet("Scenario Results")
    ws_results['A1'] = "METHODOLOGICAL SENSITIVITY ANALYSIS"
    ws_results['A1'].font = Font(bold=True, size=12)
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_results.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 3:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            if c_idx in [10, 11] and r_idx > 3:
                cell.number_format = '0.00'
            elif c_idx in [13, 14] and r_idx > 3:
                cell.number_format = '0.0'
    
    for col in ws_results.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_results.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
    
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = "KEY FINDINGS"
    ws_summary['A1'].font = Font(bold=True, size=14)
    
    cost_var = ((df['MFSP_$/GGE'].max() / df['MFSP_$/GGE'].min() - 1) * 100)
    ghg_var = ((df['GHG_gCO2e/MJ'].max() / df['GHG_gCO2e/MJ'].min() - 1) * 100)
    harmonized_mask = df['Category'].isin(['Baseline', 'Economics', 'Environmental', 'Operational', 'Combined'])

    before_cost = df['MFSP_$/GGE']
    after_cost = df[harmonized_mask]['MFSP_$/GGE']
    before_ghg = df['GHG_gCO2e/MJ']
    after_ghg = df[harmonized_mask]['GHG_gCO2e/MJ']

    def _coefficient_of_variation(series):
        return (series.std() / series.mean() * 100) if series.mean() else 0

    cv_before_cost = _coefficient_of_variation(before_cost)
    cv_after_cost = _coefficient_of_variation(after_cost) if not after_cost.empty else 0
    cv_before_ghg = _coefficient_of_variation(before_ghg)
    cv_after_ghg = _coefficient_of_variation(after_ghg) if not after_ghg.empty else 0

    reduction_cost = ((cv_before_cost - cv_after_cost) / cv_before_cost * 100) if cv_before_cost else 0
    reduction_ghg = ((cv_before_ghg - cv_after_ghg) / cv_before_ghg * 100) if cv_before_ghg else 0
    
    summary = [
        ["Metric", "Value"],
        ["Cost Range", f"${df['MFSP_$/GGE'].min():.2f} - ${df['MFSP_$/GGE'].max():.2f}/GGE"],
        ["Cost Variation", f"{cost_var:.0f}%"],
        ["GHG Range", f"{df['GHG_gCO2e/MJ'].min():.1f} - {df['GHG_gCO2e/MJ'].max():.1f} gCO2e/MJ"],
        ["GHG Variation", f"{ghg_var:.0f}%"],
        ["Cost Variance Reduction", f"{reduction_cost:.0f}%"],
        ["GHG Variance Reduction", f"{reduction_ghg:.0f}%"],
    ]
    
    for row_idx, row_data in enumerate(summary, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            ws_summary.cell(row=row_idx, column=col_idx, value=value)
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 30
    
    return wb

if __name__ == "__main__":
    print("="*70)
    print("SAF HARMONIZATION FRAMEWORK - COMPUTATIONAL ANALYSIS")
    print("="*70)
    
    print("\nRunning 14-scenario analysis...")
    df = run_full_analysis()
    
    print("Generating figures...")
    create_main_results_plot(df)
    create_harmonization_plot(df)
    
    print("Creating Excel workbook...")
    wb = create_excel_workbook(df)
    wb.save('outputs/SAF_Harmonization_Analysis.xlsx')
#%% 
    print(f"\n{'='*70}")
    print("ANALYSIS COMPLETE")
    print(f"{'='*70}")
    print(f"\nCost range: ${df['MFSP_$/GGE'].min():.2f} - ${df['MFSP_$/GGE'].max():.2f}/GGE")
    print(f"Cost variation: {((df['MFSP_$/GGE'].max() / df['MFSP_$/GGE'].min() - 1) * 100):.0f}%")
    print(f"\nGHG range: {df['GHG_gCO2e/MJ'].min():.1f} - {df['GHG_gCO2e/MJ'].max():.1f} gCO2e/MJ")
    print(f"GHG variation: {((df['GHG_gCO2e/MJ'].max() / df['GHG_gCO2e/MJ'].min() - 1) * 100):.0f}%")
    print(f"\nFiles saved in outputs/ directory")
    print(f"{'='*70}\n")
#%%